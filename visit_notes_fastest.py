from selenium import webdriver
import sys
from time import sleep
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium import webdriver
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver import ChromeOptions
from datetime import datetime
from datetime import date, timedelta
from dateutil.parser import parse
import time
from time import sleep
from selenium.common.exceptions import NoSuchElementException
from selenium.common.exceptions import StaleElementReferenceException
from selenium.common.exceptions import InvalidSessionIdException
from selenium.common.exceptions import NoSuchElementException
from selenium.common.exceptions import StaleElementReferenceException
from selenium.webdriver.support.select import Select
import os
from os import listdir
from os.path import isfile, join
import csv
from bs4 import BeautifulSoup
from bs4.element import Comment
import re
from dateutil import parser
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
from openpyxl.styles.colors import Color
from openpyxl.styles import Font
import winsound
import psycopg2

# Connect to your postgres DB
con = psycopg2.connect(
    host='localhost',
    user='postgres',
    password='winwin',
)

# Open a cursor to perform database operations
cur = con.cursor()


excel_location = 'C:/Users/johnc/Desktop/Visit note QUALITY ASSURANCE HOSPICE/visit notes qa.xlsx'
wb = load_workbook(excel_location)
ws = wb.active
max_row = ws.max_row


# this allows muliple arguments to be passed without being defined.
def p(*args):
    running = False
    if running:
        print(*args)
        # print("Len:",len(args))
        # print("Type:",type(args))
        running = False


startTime = datetime.now()
today = date.today()

# rasmussen "5013p"  Done[1190,1157,1193,villarini,lewellyn
pt_name = "Vito"  # "Ramirez" has lots.. so works fine.. gotta get it worked despite len
# ana.doneza@gmail.com   Jcarr@artofhospicecare.com
login_name = "Jcarr@artofhospicecare.com"
password = "Aohc1234!!"  # Luvzit_02!    Aohc1234!!
findings = []
final_findings = []
# disables browser notifications  # OLD ONE BELOW
options = webdriver.ChromeOptions()
all_data = []

### MAKE IT SO IT JUST COLLECTS ALL THE DATA FIRST>> THEN I PARSE IT OFFLINE from Excel on openpy ###

### THEN IT WILL GO BACK IN AND MAKE THE CHANGES afterwards   ###

options.add_experimental_option('prefs', {
    "profile.default_content_setting_values.notifications": 2,
    # Change default directory for downloads
    "download.default_directory": "C:\\Users\\johnc\\Desktop\\HPDOWNLOADS",
    "download.prompt_for_download": False,  # To auto download the file
    # just click it and it will download it to this path for pdf's at least.
    "download.directory_upgrade": True,
    # It will not show PDF directly in chrome
    "plugins.always_open_pdf_externally": True,
})

driver = webdriver.Chrome(
    "C:/Users/johnc/chromedriver.exe", options=options)

actions = ActionChains(driver)


def sound():
    for i in range(2):
        winsound.PlaySound('sound.wav', winsound.SND_FILENAME)


def alert_sound():
    for i in range(3):
        winsound.PlaySound("SystemExit", winsound.SND_ALIAS)


def perform_actions():
    """ Perform and reset actions """
    actions.perform()
    actions.reset_actions()
    for device in actions.w3c_actions.devices:
        device.clear_actions()


def find(self):

    WebDriverWait(driver, 60).until(
        EC.element_to_be_clickable((By.XPATH, self)))
    return driver.find_element(By.XPATH, self)


def load_page():
    url = driver.current_url
    driver.get(url)


def get_checked(xpath):  # don't put try statement on it because it wont go to correct one
    # sleep(.3) #1 second worked.. trying .3.  ITs needed
    WebDriverWait(driver, 30).until(
        EC.presence_of_element_located((By.XPATH, xpath)))
    item = driver.find_element(By.XPATH, xpath)
    checked = item.get_attribute('checked')
    return checked  # returns true if checked


def find_allow_page_full_load(self):
    WebDriverWait(driver, 30).until(
        EC.element_to_be_clickable((By.XPATH, self)))
    return driver.find_element(By.XPATH, self)


def find_elements(xpath):
    WebDriverWait(driver, 60).until(
        EC.presence_of_all_elements_located((By.XPATH, xpath)))
    return driver.find_elements(By.XPATH, xpath)


def find_wait(self, wait):
    WebDriverWait(driver, wait).until(
        EC.element_to_be_clickable((By.XPATH, self)))
    return driver.find_element(By.XPATH, self)


def find_stale_element(self):
    p("Looking for stale element:", self)
    ignored_exceptions = (NoSuchElementException,
                          StaleElementReferenceException,)
    your_element = WebDriverWait(driver, 30, ignored_exceptions=ignored_exceptions)\
        .until(EC.presence_of_element_located((By.XPATH, self)))
    return driver.find_element(By.XPATH, self)


def move_to_element(xpath):
    ActionChains(driver).move_to_element(xpath)
    perform_actions()


click_list = []


def click_old(self, explicit_wait_seconds):
    # sleep(10)
    try:
        var = WebDriverWait(driver, explicit_wait_seconds).until(
            EC.element_to_be_clickable((By.XPATH, self)))
        # added the click here.
        ActionChains(driver).move_to_element(var).click()
        perform_actions()
        click_list.append("0")
        return driver.find_element(By.XPATH, self).click()
    except:
        try:
            sleep(1)  # 2 worked
            driver.find_element(By.XPATH, self).click()
            sleep(1)
            print('Part 1 click', self)
            click_list.append("1")
        except:
            try:
                driver.find_element(By.XPATH, self).click()
                print('Part 2 click', self)
                click_list.append("2")
            except:
                try:
                    element = find(self)
                    ActionChains(driver).move_to_element(element).click()
                    perform_actions()
                    print("part 3 click")
                    click_list.append("3")

                except Exception as e:
                    print("Find and click did not work:", self)
                    print("EXCEPTION: ", e)


def click(self):
    try:
        element = find(self)
        actions.move_to_element(element)

        actions.click(element)
        perform_actions()
        # print("first click worked")

    except:
        # print('def click not found on xplicit wait part 1', self)
        try:
            # 2 worked  #1 usually works but maybe not if internet is slow.
            sleep(2)
            driver.find_element(By.XPATH, self).click()
            sleep(1)
            # print("Sleeping on click")
        except:
            # print('def click trying to find xpath after a sleep', self)
            try:
                driver.find_element(By.XPATH, self).click()

            except:
                try:
                    element = find(self)
                    ActionChains(driver).move_to_element(element).click()
                    perform_actions()
                    print("CLICKED IT WORK")

                except Exception as e:
                    p("Find and click did not work:", self)
                    p("EXCEPTION: ", e)


def typing(xpath_to_type_in, num_of_backspaces, words_to_type):
    try:
        sleep(.2)  # was .5
        click(xpath_to_type_in)
        var = find(xpath_to_type_in)

        for i in range(num_of_backspaces):
            var.send_keys(Keys.BACKSPACE)

        var.send_keys(words_to_type)
        sleep(.5)

    except:
        p('exception typing sleep 1')
        pass
        try:
            click(xpath_to_type_in)
            sleep(1)
            # var = driver.find_element_by_xpath(xpath_to_type_in)
            var = driver.find_element(By.XPATH, xpath_to_type_in)
            var.send_keys(words_to_type)
            sleep(1)
        except:
            p('exception typing sleep 2')
            pass
            try:
                # click(xpath_to_type_in)
                sleep(2)
                p('a')
                var = find(xpath_to_type_in)
                var.send_keys(words_to_type)
                sleep(2)
            except:
                seq = driver.find_elements_by_tag_name('iframe')
                p("NUmbe of iframes:", len(seq))
                p('exception typing sleep 3: probably didnt click link before it or iframe changed. might need to inspect and search iframe to input that xpath into switch_iframe_function')
                sys.exit()


def typing_backarrow(xpath_to_type_in, num_of_right_arrows, words_to_type):
    sleep(.2)  # was .5
    click(xpath_to_type_in)
    var = find(xpath_to_type_in)

    for i in range(num_of_right_arrows):
        sleep(.1)
        var.send_keys(Keys.ARROW_RIGHT)

    for i in range(num_of_right_arrows*2):
        sleep(.1)
        var.send_keys(Keys.BACKSPACE)

    var.send_keys(words_to_type)
    sleep(.1)


def switch_to_new_window():
    p(driver.current_window_handle)
    p(driver.window_handles)

    try:
        # sleep(4)  # 2 worked but not all the time. #trying to remove this first sleep
        second_window = driver.window_handles[1]
        sleep(3)  # worked at 4
        driver.switch_to.window(second_window)
        sleep(3)  # works better slower, added this extra wait

    except:
        # p(driver.current_window_handle)
        # p('unable to switch to window after attempt 1')
        try:
            sleep(2)
            second_window = driver.window_handles[1]
            sleep(2)
            driver.switch_to.window(second_window)
            # sleep(4)
            p(driver.current_window_handle)
            p(driver.window_handles)
        except:

            pass
            try:
                # sleep(5)
                second_window = driver.window_handles[1]
                sleep(5)  # workd at 10
                driver.switch_to.window(second_window)
                sleep(5)

            except Exception as e:
                p("Switching to new 8687 window")
                p("EXCEPTION: ", e)
                p("Played sound")
                winsound.PlaySound("SystemExit", winsound.SND_ALIAS)


def switch_to_old_window():
    try:
        # sleep(2)
        window_before = driver.window_handles[0]
        sleep(2)
        driver.switch_to.window(window_before)
        sleep(2)
    except:
        print("step22")
        pass
        try:
            sleep(4)
            window_before = driver.window_handles[0]
            sleep(4)
            driver.switch_to.window(window_before)
            sleep(4)
        except:
            p("Step33")
            pass
            try:
                # sleep(8)
                window_before = driver.window_handles[0]
                sleep(5)  # worked at 8
                driver.switch_to.windoww(window_before)
                sleep(5)
            except Exception as e:
                p("Switching to old98 window")
                p("EXCEPTION: ", e)
                p("Played sound")
                winsound.PlaySound("SystemExit", winsound.SND_ALIAS)


def back(times):
    for i in range(times):
        driver.execute_script("window.history.go(-1)")
        sleep(.2)


def get_value(xpath):  # will come back 1 for one selection or another number for another one
    WebDriverWait(driver, 60).until(
        EC.visibility_of_element_located((By.XPATH, xpath)))
    item = driver.find_element(By.XPATH, xpath)
    return item.get_attribute('value')


def switch_iframe(xpath_to_iframe):
    seq = driver.find_elements_by_tag_name('iframe')
    p("NUmbe of iframes:", len(seq))
    iframe_xp = xpath_to_iframe
    iframe = find(iframe_xp)
    driver.switch_to.frame(iframe)
    p("USe this to close iframe and go back to original: driver.switch_to.default_content()")
    # then you have to close it


def check_bp_range(bp):
    vs2 = bp.replace('/', "")
    diastolic = int(vs2[-2:])
    systolic = int(vs2[:-2])
    value = ''

    if systolic > 160:
        return "hypertensive"
        if systolic > 220:
            value = "hyper hyper Systolic bps " + bp

    elif systolic < 90:
        value = "hypotensive"
        if systolic < 40:
            value = "extra hypo systolic " + bp

    if diastolic > 100:
        value = "hypertensive"
        if diastolic > 180:
            value = "Very high diastolic " + bp

    elif diastolic < 50:
        value = "hypotensive"
        if diastolic < 35:
            value = "extra hypo diastolic " + bp

    return value


def get_all_values(xpath, printname):
    print(printname)
    data_dict = {}
    WebDriverWait(driver, 40).until(
        EC.visibility_of_element_located((By.XPATH, xpath)))
    item = driver.find_element(By.XPATH, xpath)

    keyname = printname

    outer_html = item.get_attribute('outerHTML')
    # print(outer_html)

    if 'type="radio"' in outer_html:
        data = item.get_attribute('checked')

        print("CHECKED1")
        print(data)
        data_dict[keyname] = data
        return data_dict

    if 'class="DrpItems10"' in outer_html:
        if 'select' in outer_html:
            try:
                select = Select(driver.find_element(
                    By.XPATH, xpath))  # does find work?
                selected_option = select.first_selected_option
                data = selected_option.text
                print("FOUND SELECTED1", data)
                data_dict[keyname] = data
                return data_dict
            except:
                print("passed dropdown")

    data = item.get_attribute('value')
    print("FOUND VALUE value")
    print(data)

    if data == 'on':
        data = item.get_attribute('checked')

        print("CHECKED2")
        print(data)
        data_dict[keyname] = data
        return data_dict

    if data == '':
        print("passing empty str")

    else:
        print("TRYING TO APPEND IT...", keyname)
        data_dict[keyname] = data
        return data_dict

    data = item.get_attribute('checked')

    print("CHECKEDc")
    print(data)

    data = item.text
    # print("FOUND TEXT")

    print(printname)
    # print("data2",data)

    data_dict[keyname] = data

    # print(" DONE  DONE DONE DONE DONE DONE ")
    print("DATA_DICT: ", data_dict)

    return data_dict


def get_selected(xpath, print_name):
    try:
        WebDriverWait(driver, 30).until(
            EC.visibility_of_element_located((By.XPATH, xpath)))
        select = Select(driver.find_element(
            By.XPATH, xpath))  # does find work?
        selected_option = select.first_selected_option
        selected = selected_option.text
        # print(print_name, "SELECTED ITEM:", selected)
        return selected
    except:
        try:
            stale_finder = find_stale_element(xpath)
            selected_option = stale_finder.first_selected_option
            selected = selected_option.text
        except:
            print("Trying to get value exception since get_selected failed")
            value = get_value(xpath)
            return value


def get_lock_status():
    lock_status = ''
    status = find('//*[@id="ctl00_ContentPlaceHolder1_ImgProgressStatus"]')
    statusbar = status.get_attribute("title")
    print("STATUSBAR TITLE: ", statusbar)
    if 'Electronically Signed,' in statusbar:
        print("ITS LOCKED")
        lock_status = 'locked'
    else:
        lock_status = 'unlocked'

    return lock_status

# must use variable lock_status = unlock_note()


def unlock_note():
    print("------------unlocking-=-===")
    lock_status = get_lock_status()
    if lock_status == 'locked':
        print("-----------UNLOCKING NOTE------------------")
        unlock = '//*[@id="ctl00_ContentPlaceHolder1_btnUnlockNew"]'
        print("UNLOCKING")
        click(unlock)
        try:
            WebDriverWait(driver, 7).until(EC.alert_is_present(),
                                           'Timed out waiting for ' +
                                           'confirmation popup to appear.')
            alert = driver.switch_to.alert
            alert.accept()
            print("alert accepted")
        except:
            pass

    sleep(6)

    actions.send_keys(Keys.ENTER)
    perform_actions()
    # print("actions done")
    sleep(2)

    # find('//*[@id="ctl00_ContentPlaceHolder1_ImgExpandAll"]')
    click('//*[@id="ctl00_ContentPlaceHolder1_ImgExpandAll"]')
    # print("Expanded")
    sleep(6)


repaired = ''


def lock_note():
    print("--------------lock_note_________________")
    lock_status = get_lock_status()
    if lock_status == "locked":
        try:
            print("Lockin it")
            lock = '//*[@id="ctl00_ContentPlaceHolder1_btnlockNew"]'
            find_wait(lock, 2)
            click(lock)
            sleep(4)
            actions.send_keys(Keys.ENTER)
            perform_actions()
        except:
            print("didn't need to lock")

    save_button = '//*[@id="ctl00_ContentPlaceHolder1_btnSave0"]'
    for i in range(2):
        click(save_button)
        sleep(.2)
    print("Pressed Save sleep4")
    sleep(4)

    repaired = 'yes'
    return repaired


def click_note(xp):
    print("--------------click note------------")
    unlock_note()
    click(xp)
    print("CLICKING NOTE")
    lock_note()
    repaired = 'yes'
    return repaired
    # sleep(30)


def type_note(xp, print_name):
    unlock_note()
    typing(xp, 8, print_name)
    lock_note()
    print("TYPED NOTE DID IT WORK ? SLEEPING 30")
    repaired = 'yes'
    return repaired
    # sleep(30)


# make sure it wasn't expanded first if it was you gotta add that expanded url as url
def write_page_file(file_name):
    # to use with bs4
    print('Getting page source')
    url = driver.current_url
    driver.get(url)
    page_html = driver.page_source

    soup = BeautifulSoup(page_html, 'html.parser')
    with open(file_name, 'wt', encoding='utf-8') as html_file:  # writes code name
        for line in soup.prettify():
            html_file.write(line)
        print(file_name, "Written to file.")


def click_idg():
    click('//*[@id="ctl00_TreeView1t15"]')


def click_visit_note_tab():
    click('//*[@id="ctl00_TreeView1t30"]')


def append_to_excel(list_to_add_in_new_row):
    ws.append(list_to_add_in_new_row)
    wb.save(excel_location)


# 10 is good for excel
def add_char_every_certain_num_char_in_string(how_often_num, sentence):
    sentence2 = sentence.split()
    words_length = 0
    word_index = 0
    for word in sentence2:
        length = len(word)
        words_length += length
        word_index += 1
        # print("WORDS LENGTH:",words_length)
        if words_length > how_often_num:
            sentence2.insert(word_index, '\n')
            words_length = 0
        else:
            continue

    finished = ' '.join(sentence2)
    # print(finished)
    return finished


# REQUIRES YOU TO HAVE CURRENT ROW BEFORE
def input_list_to_newline_excel(column, row, list_name):
    # counts the row to get the current row to put findings in right spot
    for row in ws.iter_rows(min_row=1, max_row=max_row):
        iterated_row = str(row[0].row)
        current_row = iterated_row


def press_enter():
    actions.send_keys(Keys.ENTER)
    perform_actions()


def typing_backarrow(xpath_to_type_in, num_of_right_arrows, words_to_type):
    sleep(.2)  # was .5
    click(xpath_to_type_in)
    var = find(xpath_to_type_in)

    for i in range(num_of_right_arrows):
        sleep(.1)
        var.send_keys(Keys.ARROW_RIGHT)

    for i in range(num_of_right_arrows*2):
        sleep(.1)
        var.send_keys(Keys.BACKSPACE)

    var.send_keys(words_to_type)
    sleep(.1)


def add_dict(keyname, value):
    basic_dict = {}
    basic_dict[keyname] = value
    all_data.append(basic_dict)
    return


def expand_assessment_tab():
    expand_assessment_xp = "//a[@id='ctl00_TreeView1n7']"
    click(expand_assessment_xp)


def click_nursing():
    nursing_xp = "//*[text()='Nursing']"
    click(nursing_xp)


def open_comprehensive_assessment():
    print("opening comprehensive")
    open_comprehensive_xp = "(//*[text()='Comprehensive']/preceding-sibling::td)[1]"
    open_comp_for_pt_german = '//*[@id="ctl00_ContentPlaceHolder1_GridView1"]/tbody/tr[10]/td[1]/input'
    try:
        click(open_comprehensive_xp)
        expand_all_ass_tabs = "//*[@id='ctl00_ContentPlaceHolder1_ImgExpandAll']"
        click(expand_all_ass_tabs)

    except:
        click(open_comp_for_pt_german)
        click(expand_all_ass_tabs)


def expand_once_comp_opened():  # second one in case first one didn't work
    expand_all_ass_tabs = "//*[@id='ctl00_ContentPlaceHolder1_ImgExpandAll']"
    click(expand_all_ass_tabs)


#####################################################################################
############################## BEGINNING#############################################

def search_pt():
    search_pt = "//input[@name='ctl00$TopSearch$txtSearch']"
    typing(search_pt, 0, pt_name)
    sleep(1)
    actions.send_keys(Keys.ENTER)
    # print("pressed Enter")
    perform_actions()


def login_hospice():
    p("_---------------------LOGIN_HOSPICE----------------------")
    driver.get("https://hospicemd.com/")
    driver.maximize_window()

    login = """//input[@name="LoginIdtxtbox"]"""

    typing(login, 0, login_name)

    pw = """//input[@name="Passwordtxtbox"]"""

    typing(pw, 0, password)

    login = """//input[@name="loginbut"]"""
    click(login)

    # patient_list = "//a[@id='ctl00_HP2']"
    # click(patient_list, 8)

    search_pt()

    # print("actions performed")
login_hospice()


def find_soc():
    expand_assessment_xp = "//a[@id='ctl00_TreeView1n7']"
    click(expand_assessment_xp)

    nursing_xp = "//*[text()='Nursing']"
    click(nursing_xp)

    print("opening comprehensive")
    open_comprehensive_xp = "(//*[text()='Comprehensive']/preceding-sibling::td)[2]"
    click(open_comprehensive_xp)

    soc = find(open_comprehensive_xp)
    soc = soc.text
    print("SOC:", soc)
    add_dict('soc', soc)
    return soc

# soc = find_soc()


def find_dx():
    dx_dict = {}
    print("opening comprehensive to get dx")
    open_comprehensive_xp = "(//*[text()='Comprehensive']/preceding-sibling::td)[1]"
    open_comp_for_pt_german = '//*[@id="ctl00_ContentPlaceHolder1_GridView1"]/tbody/tr[10]/td[1]/input'
    try:
        click(open_comprehensive_xp)
        expand_all_ass_tabs = "//*[@id='ctl00_ContentPlaceHolder1_ImgExpandAll']"
        click(expand_all_ass_tabs)

    except:
        click(open_comp_for_pt_german)
        click(expand_all_ass_tabs)

    expand_all_ass_tabs = "//*[@id='ctl00_ContentPlaceHolder1_ImgExpandAll']"
    click(expand_all_ass_tabs)
    dx = get_selected(
        '//*[@id="ctl00_ContentPlaceHolder1_DrpPDDisease0"]', 'Diagnosis')
    add_dict('dx', dx)
    return dx
# diagnosis = find_dx()


def check_missing_visits(start_date):
    print("CHECKING MISSING VISITS")
    reports = click('//*[@id="ctl00_HyperLink1"]')
    missing_visits = click('//*[@id="ctl00_ContentPlaceHolder1_MVisits"]')
    sleep(1)

    from_ = typing_backarrow(
        '//*[@id="ctl00_ContentPlaceHolder1_txtfrom"]', 20, start_date)
    sleep(1)
    submit = click('//*[@id="ctl00_ContentPlaceHolder1_btnSearch"]')
    start_date_rows = '/html/body/form/table/tbody/tr[1]/td/div/table/tbody/tr[2]/td/table/tbody/tr[3]/td[2]/div/table/tbody/tr/td/div/div[1]/div/div[*]/table/tbody/tr/td[2]/span'
    try:
        start_dates = find_elements(start_date_rows)
    except:
        print("No missing visits found")
        return  # because if their are none it will skip function

    rows_num = len(start_dates)

    names_num = 2
    iterate_num = 1
    expand_num = 1  # this one goes up +3 each time since element isn't read on every number
    for i in range(rows_num):
        expander = f'/html/body/form/table/tbody/tr[1]/td/div/table/tbody/tr[2]/td/table/tbody/tr[3]/td[2]/div/table/tbody/tr/td/div/div[1]/div/div[{expand_num}]/table/tbody/tr/td[15]/img'
        click(expander)
        expand_num += 3
        sleep(.3)

    # FIND NAME

    ### TRYING TO SEE IF I CAN PRINT NAMES AND DATES ##
    all_rows = find_elements(
        '/html/body/form/table/tbody/tr[1]/td/div/table/tbody/tr[2]/td/table/tbody/tr[3]/td[2]/div/table/tbody/tr/td/div/div[1]/div/div')
    for row in all_rows:
        row_text = row.text

        potential_date_row = row_text[:3]
        # print("potential_date_row",potential_date_row)

        try:
            potential_date_row = int(potential_date_row)
            dates = row_text
            # print("DATES:",dates)
        except:
            pass

        if pt_name in row_text:
            findings.append("Missing Visit Notes" + dates[4:-13])
            print("FOUND MISSING NOTES AND ADDED TO FINDING")

    # ########################### END #################

    # name_row = '/html/body/form/table/tbody/tr[1]/td/div/table/tbody/tr[2]/td/table/tbody/tr[3]/td[2]/div/table/tbody/tr/td/div/div[1]/div/div[*]/div/div[*]/div/table/tbody/tr'
    # name_rows = find_elements(name_row)
    # # print("Name_row LEN:",len(name_rows))

    # ## TESTING TO SEE IF I GET DATES AND ROW DATA
    # dates_rows = '/html/body/form/table/tbody/tr[1]/td/div/table/tbody/tr[2]/td/table/tbody/tr[3]/td[2]/div/table/tbody/tr/td/div/div[1]/div/div'
    # date_row = find_elements(dates_rows)

    # row_list = []
    # for row in date_row:
    #     row = row.text
    #     row_list.append(row)

    # # save to list then parse list
    # # print("ROW_LIST = ",row_list)

    # missing_note_list = []

    # ######################################## PASTED IN

    # row_num = 0
    # for row in date_row:
    #     try:
    #         finder = row.index(pt_name)
    #         # print('finder',finder)
    #         # print('row num',row_num)

    #         date = date_row[row_num-1]
    #         # print('date',date)

    #     except:
    #         pass

    #     row_num += 1

    # row_num = 0
    # miss_num = 1
    # for row in name_rows:
    #     row = row.text      ##GOTTA CUT THIS STRING
    #     # print("ROW:",row)

    #     if pt_name in row:
    #         miss_num += 1
    #         print("MISSING NOTE ",row)
    #         findings.append("MISSING NOTES"+ row)
    #         missed = 'missing note'
    #         add_dict(missed,row)
    #         row_num += 1
    # else:
    #     pass

    # print(row_num)

# check_missing_visits(soc)


def click_most_recent_assessment():
    print("-------------Gather comp baseline -------------")
    search_pt()
    expand_assessment_tab()
    click_nursing()

    # GO TO COMP or recert if there is recert

    # PICK MOST RECENT RECERT IF THERE IS ONE>
    try:
        assessment = click(
            '//*[@id="ctl00_ContentPlaceHolder1_GridView1"]/tbody/tr[2]/td[1]/input')
        expand_once_comp_opened()
    except:
        findings.append("No Nurse Assessment Found")
        print("No assessment found")

        # HAVE IT CONV ASS TO DATETIME IF I NEED IT TO THEN OPEN NEWEST INFO
click_most_recent_assessment()

# gather_nurse_ass needs click_most_recent_active


def gather_nurse_assessment_data():

    nurse_assessment_list = []

    assessment_xp_dict = {
        'visit_date': '//*[@id="ctl00_ContentPlaceHolder1_txtHDate"]',
        'in_person': '//*[@id="ctl00_ContentPlaceHolder1_Assesscodeid_0"]',
        'recert': '//*[@id="ctl00_ContentPlaceHolder1_AssessmentType_2"]',
        'staff_assigned': '//*[@id="ctl00_ContentPlaceHolder1_DropStaffAssigned"]',
        'covid_screen_yes': '//*[@id="ctl00_ContentPlaceHolder1_AssessmentType_2"]',
        'pt_report_pos': '//*[@id="ctl00_ContentPlaceHolder1_rbPositive_0"]',
        'lose_taste': '//*[@id="ctl00_ContentPlaceHolder1_rbCoronaVirus_0"]',
        'addit_concerns': '//*[@id="ctl00_ContentPlaceHolder1_rbPatientPcgVisit_0"]',
        'covid_followed': '//*[@id="ctl00_ContentPlaceHolder1_chkCoronaInstructions"]',
        'utilized_ppe': '//*[@id="ctl00_ContentPlaceHolder1_chkPpe"]',

        ### Pain Screen ###
        'pt_verbal_pain': '//*[@id="ctl00_ContentPlaceHolder1_RD_PN_PAL_0"]',
        'pt_uncomfortable_no': '//*[@id="ctl00_ContentPlaceHolder1_RD_PN_UN_0"]',

        ### VITALS ###
        'temp': '//*[@id="ctl00_ContentPlaceHolder1_Vital_temp"]',
        'pulse': '//*[@id="ctl00_ContentPlaceHolder1_Vital_Pulse"]',
        'resp': '//*[@id="ctl00_ContentPlaceHolder1_Vital_Resp"]',
        'bp': '//*[@id="ctl00_ContentPlaceHolder1_Vital_BP"]',
        'weight': '//*[@id="ctl00_ContentPlaceHolder1_Vital_Wt"]',
        'mac': '//*[@id="ctl00_ContentPlaceHolder1_Vital_Mac"]',
        'o2_sat': '//*[@id="ctl00_ContentPlaceHolder1_Vital_O2SAT"]',

        ### MOBILITY ###
        'ambulatory': '//*[@id="ctl00_ContentPlaceHolder1_RAD_CEMO_AN"]',
        'ambu_max_assist': '//*[@id="ctl00_ContentPlaceHolder1_Chk_CEMO_MX"]',

        'non_ambulatory': '//*[@id="ctl00_ContentPlaceHolder1_RAD_CEMO_AN1"]',
        'bedbound': '//*[@id="ctl00_ContentPlaceHolder1_ChK_CEMO_BB"]',
        'non_ambu_max_assist': '//*[@id="ctl00_ContentPlaceHolder1_Chk_CEMO_MXA"]',

        ### ADL ASSESSMENT ###
        'ambulation': '//*[@id="ctl00_ContentPlaceHolder1_DRP_CEAD_A"]',
        'toileting': '//*[@id="ctl00_ContentPlaceHolder1_DRP_CEAD_C"]',
        'transfer': '//*[@id="ctl00_ContentPlaceHolder1_DRP_CEAD_T"]',
        'dressing': '//*[@id="ctl00_ContentPlaceHolder1_DRP_CEAD_D"]',
        'feeding': '//*[@id="ctl00_ContentPlaceHolder1_DRP_CEAD_F"]',
        'bathing': '//*[@id="ctl00_ContentPlaceHolder1_DRP_CEAD_B"]',

        ### UPDATE CHANGE IN DIAGNOSIS ###
        'primary_dx': '//*[@id="ctl00_ContentPlaceHolder1_PrimaryDx"]',
        'primary_disease': '//*[@id="ctl00_ContentPlaceHolder1_DrpPDDisease0"]',

        ### KPS PPS FAST NYHA ###
        'kps': '//*[@id="ctl00_ContentPlaceHolder1_KPSList"]',
        'pps': '//*[@id="ctl00_ContentPlaceHolder1_PPSList"]',
        'fast': '//*[@id="ctl00_ContentPlaceHolder1_FASTList"]',

        ### NATURE # CONDITION OF TERMINAL ILLNESS /LCD Eligibility ###
        'kps_less_70': '//*[@id="ctl00_ContentPlaceHolder1_Dem_7"]',
        'pps_less_70': '//*[@id="ctl00_ContentPlaceHolder1_Dem_8"]',
        'fast_greater_7a': '//*[@id="ctl00_ContentPlaceHolder1_Dem_9"]',
        'unable_ambulate': '//*[@id="ctl00_ContentPlaceHolder1_Dem_10"]',
        'unable_dress': '//*[@id="ctl00_ContentPlaceHolder1_Dem_11"]',
        'unable_bath': '//*[@id="ctl00_ContentPlaceHolder1_Dem_12"]',
        'urin_fec_incont': '//*[@id="ctl00_ContentPlaceHolder1_Dem_13"]',
        'unable_speak_six': '//*[@id="ctl00_ContentPlaceHolder1_Dem_14"]',
        'aspiration_pneu': '//*[@id="ctl00_ContentPlaceHolder1_NC_PDrp"]',
        'pyenophritis': '//*[@id="ctl00_ContentPlaceHolder1_Dem_16"]',
        'septicemia': '//*[@id="ctl00_ContentPlaceHolder1_NC_SDrp"]',
        'decubitis': '//*[@id="ctl00_ContentPlaceHolder1_NC_UDrp"]',
        'recurrent_fever': '//*[@id="ctl00_ContentPlaceHolder1_NC_RDrp"]',
        'weight_loss': '//*[@id="ctl00_ContentPlaceHolder1_NC_WDrp"]',

        ### BODY SYSTEM ###
                           ## SYMPTOMS ##
        'peaceful': '//*[@id="ctl00_ContentPlaceHolder1_Nue_DEMpe"]',
        'anxious': '//*[@id="ctl00_ContentPlaceHolder1_Nue_DEMan"]',
        'confused': '//*[@id="ctl00_ContentPlaceHolder1_Nue_DEMco"]',
        'angry': '//*[@id="ctl00_ContentPlaceHolder1_Nue_DEMag"]',
        'restless': '//*[@id="ctl00_ContentPlaceHolder1_Nue_DEMre"]',
        'agitated': '//*[@id="ctl00_ContentPlaceHolder1_Nue_DEMat"]',
        'depressed': '//*[@id="ctl00_ContentPlaceHolder1_Nue_DEMot"]',
        'seizure': '//*[@id="ctl00_ContentPlaceHolder1_Nue_OTHse"]',
        'combative': '//*[@id="ctl00_ContentPlaceHolder1_Nue_OTHco"]',
        'sundowning': '//*[@id="ctl00_ContentPlaceHolder1_Nue_OTHsu"]',
        'tremors': '//*[@id="ctl00_ContentPlaceHolder1_Nue_OTHtr"]',
        'other_symptoms': '//*[@id="ctl00_ContentPlaceHolder1_Nue_OTHotText"]',

        # LEVEL OF CONSCIOUSNESS:
        'awake': '//*[@id="ctl00_ContentPlaceHolder1_Nue_LOCaw"]',
        'alert': '//*[@id="ctl00_ContentPlaceHolder1_Nue_LOCal"]',
        'lethargic': '//*[@id="ctl00_ContentPlaceHolder1_Nue_LOCle"]',
        'min_responsive': '//*[@id="ctl00_ContentPlaceHolder1_Nue_LOCmr"]',
        'coma': '//*[@id="ctl00_ContentPlaceHolder1_Nue_LOCco"]',

        ### ORIENTED TO ###
        'time_oriented': '//*[@id="ctl00_ContentPlaceHolder1_Nue_ORTtm"]',
        'place_oriented': '//*[@id="ctl00_ContentPlaceHolder1_Nue_ORTpl"]',
        'person_oriented': '//*[@id="ctl00_ContentPlaceHolder1_Nue_ORTpr"]',
        'disoriented': '//*[@id="ctl00_ContentPlaceHolder1_Nue_ORTds"]',

                ### PSYCH HX ###
        'no_psych_hx': '//*[@id="ctl00_ContentPlaceHolder1_Nue_PHXno"]',
        'bipolar_hx': '//*[@id="ctl00_ContentPlaceHolder1_Nue_PHXbi"]',
        'ocd_hx': '//*[@id="ctl00_ContentPlaceHolder1_Nue_PHXoc"]',
        'schizo_hx': '//*[@id="ctl00_ContentPlaceHolder1_Nue_PHXsc"]',
        'depression_hx': '//*[@id="ctl00_ContentPlaceHolder1_Nue_PHXde"]',

                ### Communication voice speech ###
        'speech_normal': '//*[@id="ctl00_ContentPlaceHolder1_Nue_CVPno"]',
        'aphasia': '//*[@id="ctl00_ContentPlaceHolder1_Nue_CVPap"]',
        'slurred_speech': '//*[@id="ctl00_ContentPlaceHolder1_Nue_CVPsl"]',
        'other_speech': '//*[@id="ctl00_ContentPlaceHolder1_Nue_CVPotText"]',
        'limited_six_words': '//*[@id="ctl00_ContentPlaceHolder1_Nue_SS"]',

                ## SENSORY DEFICIT ##
        'blind': '//*[@id="ctl00_ContentPlaceHolder1_CHK_CE_BD"]',
        'hard_hearing': '//*[@id="ctl00_ContentPlaceHolder1_CHK_CE_HH"]',

                ## BALANCE ##
        'normal_balance': '//*[@id="ctl00_ContentPlaceHolder1_Nue_Balni_0"]',
        'balance_impaired': '//*[@id="ctl00_ContentPlaceHolder1_Nue_Balni_1"]',

        ### CARDIOVASCULAR ###
        'normal_bp': '//*[@id="ctl00_ContentPlaceHolder1_BPLevel_0"]',
        'hypotension': '//*[@id="ctl00_ContentPlaceHolder1_BPLevel_1"]',
        'hypertension': '//*[@id="ctl00_ContentPlaceHolder1_BPLevel_2"]',

        #Pulse#
        'apical_regular': '//*[@id="ctl00_ContentPlaceHolder1_Car_PULar"]',
        'apical_irreg': '//*[@id="ctl00_ContentPlaceHolder1_Car_PULai"]',
        'apical_weak': '//*[@id="ctl00_ContentPlaceHolder1_Car_PULaw"]',
        'apical_tachy': '//*[@id="ctl00_ContentPlaceHolder1_Car_PULat"]'


        # KEEP ADDING ALL COMPREHENSIVE XPS HERE













    }
    # NOW JUST ADD ALL THE OTHER XPS

    for key, value in assessment_xp_dict.items():
        print(key)
        print(value)
        assessment_dict = get_all_values(value, key)
        nurse_assessment_list.append(assessment_dict)

    # WHERE SHALL I STORE ALL THIS DATA?  HOW CAN I MAKE IT FASTER

    return nurse_assessment_list


assessment_list = gather_nurse_assessment_data()

print("assessment info")

for assessment in assessment_list:
    print(assessment)


#### COLLECT COMP OR RECERT BASELINE ####
    # GET DX , PROBLEMS,
    ### PUT BASELINE INFO IN SQL ###


# COLECT PHYSICIAN ORDERS

# COLLECT IDG


# COLLECT COMM LOG

# COLLECT VISIT NOTES


# def idg_main():
#     print("IDG MAIN STARTING")
#     ####### GOING BACK HERE BECAUSE I WENT TO MISSING NOTES FIRST
#     search_pt = "//input[@name='ctl00$TopSearch$txtSearch']"
#     typing(search_pt, 0, pt_name)
#     sleep(1)
#     actions.send_keys(Keys.ENTER)
#     perform_actions()


#     click_idg()
#     print("STARTING MAIN")
#     idg_page = "idg_main_page"
#     write_page_file(idg_page)

#     patient_name = 'Patient1'
#     patient_dict = {}


#     with open(idg_page, 'r', encoding='utf-8') as file:
#         soup = BeautifulSoup(file, 'html.parser')
#         idg_table = soup.find(id=re.compile("ctl00_ContentPlaceHolder1_GridView1"))
#         rows = idg_table.findChildren('tr')

#         date_num = 0
#         ############### GETS DATES FROM THE ROWS OF THE IDG  SAVES THEM TO PATIENT_DICT ##################
#         for row in rows[1:]:
#             num_rows = len(rows)
#             row_text = row.get_text(" ", strip=True)
#             date = row_text[:8]
#             date_num += 1
#             var_name = 'idg_date' + str(date_num)
#             patient_dict[var_name]=date

#     #SAVED IDG DATES INSICE PATIENT_DICT_ITEMS AS idg_date1 iterating to all of them
#     #CONVERTING DATES to DATETIME TO SEE WHICH NOTES ARE BETWEEN THOSE DATES
#     most_recent_idg= patient_dict['idg_date1']
#     previous_idg= patient_dict['idg_date2']
#     recent_idg = parser.parse(most_recent_idg)
#     previous_idg = parser.parse(previous_idg)
#     print("RECENT IDG DATE:",recent_idg,"PREVIOUS:",previous_idg)
#     print("FINISHED IDG MAIN")
#     return recent_idg,previous_idg,patient_dict

# # idg_main()

# ################### GATHER IDG NURSE NOTES FROM ALL OF THEM ######################
# def idg_notes():
#     print("---------------------STARTING IDG-NOTES ----------------------------")
#     ##########SAVES IDG NOTES
#     write_page_file("idg_previous")
#     page_num_links = '/html/body/form/table/tbody/tr[1]/td/div/table/tbody/tr[2]/td/table/tbody/tr[4]/td[2]/div/table/tbody/tr/td[2]/div[2]/div/table/tbody/tr[17]/td/table/tbody/tr/td[*]'
#     #############    ITERATING THROUGH PAGES IN IDG ########################
#     idg_links_xps = '/html/body/form/table/tbody/tr[1]/td/div/table/tbody/tr[2]/td/table/tbody/tr[4]/td[2]/div/table/tbody/tr/td[2]/div[2]/div/table/tbody/tr[*]/td[1]/input'

#     try:
#         recent_idg = '//*[@id="ctl00_ContentPlaceHolder1_GridView1"]/tbody/tr[2]/td[1]/input'
#         found_idg = find_wait(recent_idg, 10)
#         print("FOUND RECENT IDG")

#         print("looking for idg_links")
#         idg_links = find_elements(idg_links_xps)
#         num_of_idg_links = len(idg_links)
#         print("found them",num_of_idg_links)
#     except:
#         print('NO IDG links found')


#     print("looking for one page")
#     one_page = '//*[@id="ctl00_ContentPlaceHolder1_GridView1"]/tbody/tr[17]/td/table/tbody/tr/td[1]/span'

#     try:
#         found_pages = find_wait(one_page, 10)
#     #############    ITERATING THROUGH PAGES IN IDG ########################
#         pages = find_elements(page_num_links) # added this to only look for page before trying to click below
#     except:
#         print('No pages found')

#     append_count = 0
#     link_num = 1

#     try:
#         for i in range(len(pages)):
#             print("Loop1:",i)
#             sleep(2)   #CLICKS PAGE THEN CLICKS ALL LINKS IN THEM
#             click(f'/html/body/form/table/tbody/tr[1]/td/div/table/tbody/tr[2]/td/table/tbody/tr[4]/td[2]/div/table/tbody/tr/td[2]/div[2]/div/table/tbody/tr[17]/td/table/tbody/tr/td[{link_num}]' )
#             link_num +=1
#             print("clicked first page")

#             link_counter = 2 #starting 2 since xpath starts w 2
#             date_note_list = []
#             for i in range(num_of_idg_links):
#                 print("LOOPA:",i)
#                 try:
#                     link = click(f'/html/body/form/table/tbody/tr[1]/td/div/table/tbody/tr[2]/td/table/tbody/tr[4]/td[2]/div/table/tbody/tr/td[2]/div[2]/div/table/tbody/tr[{link_counter}]/td[1]/input')
#                 except:
#                     print("BREAKING LOOP")
#                     break

#                 link_counter += 1
#                 # print("LINK_COUNTER:",link_counter)

#                 ############ CLICKS IDG LINK< OPENS GATHERS DATE NURSE NOTES AND WRITES TO EXCEL FOR ALL PAGES
#                 expand_idg_discussion = click('//*[@id="ctl00_ContentPlaceHolder1_ImgPnlComment"]')
#                 date = get_value('//*[@id="ctl00_ContentPlaceHolder1_txtHDate"]')

#                 nurse_note = find('//*[@id="ctl00_ContentPlaceHolder1_TxtComments"]')
#                 nurse_note = nurse_note.text
#                 notes_w_new_lines = add_char_every_certain_num_char_in_string(50,nurse_note)
#                 if len(nurse_note) < 5:
#                     findings.append("Nurse NOTES MISSING IDG ON "+ date)

#                 md_note = find('//*[@id="ctl00_ContentPlaceHolder1_TxtCommentsOTH"]')
#                 md_note = md_note.text
#                 md_note_new_lines = add_char_every_certain_num_char_in_string(50,md_note)

#                 if len(md_note) < 5:
#                     findings.append("MD NOTES MISSING IDG ON "+ date)
#                 # print("MD NOTES NEW LINES:",md_note_new_lines)

#                 # nurse_note = find('//*[@id="ctl00_ContentPlaceHolder1_TxtComments"]').text
#                 # date = "IDG: " + date
#                 print("IDG DATE:",date)
#                 date_note_list.append(date)

#                 #add nurse name

#                 #add patient name above, soc,dx

#                 #rn discussion notes needs to say anas note that they are looking at note
#                 #check to see if careplans are completed or not

#                 date_note_list.append(notes_w_new_lines)
#                 date_note_list.append(md_note_new_lines)

#                 #add md note

#                 print("appending to excel")

#                 append_to_excel(date_note_list)
#                 append_count += 1
#                 print("APPEND NUM A:",append_count)
#                 date_note_list = []
#                 back(1)  #have to click back to go back once.
#                 print("PRESSED BACK A")


#             idg_links = find_elements(idg_links_xps)
#             print("found new idg links")
#         print("GATHERED DATA FROM ALL IDG")
#     except:
#         print("NO pages found and or loop didn't close correctly.. need another try/except if it has a lot of pages since it doesn't end smooth")
# ######################## THIS LOOP IS IF THERE IS NO PAGE BUT HAS SOME IDG LINKS###################

#     link_counter = 2

#     for i in range(num_of_idg_links):
#         date_note_list = []
#         print("LOOP2:",i)
#         try:
#             link = click(f'/html/body/form/table/tbody/tr[1]/td/div/table/tbody/tr[2]/td/table/tbody/tr[4]/td[2]/div/table/tbody/tr/td[2]/div[2]/div/table/tbody/tr[{link_counter}]/td[1]/input')
#         except:
#             print("BREAKING LOOP")
#             break

#         link_counter += 1
#         print("LINK_COUNTER:",link_counter)

#         ############ CLICKS IDG LINK< OPENS GATHERS DATE NURSE NOTES AND WRITES TO EXCEL FOR ALL PAGES

#         #added this because it didn't open it one time.
#         try:
#             expand_idg_discussion = click('//*[@id="ctl00_ContentPlaceHolder1_ImgPnlComment"]')
#             date = get_value('//*[@id="ctl00_ContentPlaceHolder1_txtHDate"]')
#         except:
#             print("FINISHED IDG THROUGH EXCEPTION")
#             return

#         nurse_note = find('//*[@id="ctl00_ContentPlaceHolder1_TxtComments"]')
#         nurse_note = nurse_note.text
#         notes_new_lines =add_char_every_certain_num_char_in_string(50,nurse_note)
#         if len(nurse_note) < 5:
#             findings.append("Nurse NOTES MISSING IDG ON "+ date)


#         md_note = find('//*[@id="ctl00_ContentPlaceHolder1_TxtCommentsOTH"]')
#         md_note = md_note.text
#         md_note_new_lines = add_char_every_certain_num_char_in_string(50,md_note)
#         if len(md_note) < 5:
#             findings.append("MD NOTES MISSING IDG ON "+ date)

#         # print("MD NOTES NEW LINES:",md_note_new_lines)

#         print("IDG DATE:",date)


#         ##########write function
#         #get dr name:  if dr Rajpara with missing notes and signatures send to maria
#         #write note that if soc and date same of note..


#         date_note_list.append(date)
#         date_note_list.append(notes_new_lines)
#         date_note_list.append(md_note_new_lines)


#         ############## adds signatures to excel unless they are blank as "Other Signature"##########

#         signatures = find_elements('/html/body/form/table/tbody/tr[1]/td/div/table/tbody/tr[2]/td/table/tbody/tr[4]/td[2]/div/table/tbody/tr/td[2]/table[2]/tbody/tr[3]/td/div[3]/div[2]/div/span[*]/input')
#         signature_list = []
#         for signature in signatures:
#             signature = signature.get_attribute('value')
#                 #if this doesn't work then iterate through these and get value since saved under value
#             if signature == 'Other Signature':
#                 pass
#             elif signature =='MD Signature':
#                 md_string = "MD Signature Missing IDG:"+ date
#                 if md_string not in findings:
#                     findings.append(md_string)
#             elif signature == 'MSW Signature':
#                 ms_string = "MSW Signature Missing IDG:"+ date
#                 if ms_string not in findings:
#                     findings.append(ms_string)

#             else:
#                 print("SIGNED:",signature)
#                 signature_list.append(signature)

#         # findings.append(signature_list)

#         ########### CHECKS TO MAKE SURE ONE OF THESE 4 THINGS ARE IN SIGNATURES
#         sign_check_list = ["MD","RN","SC","SW"]
#         signatures_string = " ".join(signature_list)
#         missing_list = []
#         print("SIGN STRING",signatures_string)
#         for abbreviation in sign_check_list:
#             if abbreviation in signatures_string:
#                 pass
#             else:
#                 missing_list.append(abbreviation)
#                 print("did not find abbreviation in signature",abbreviation)

#         message_list = []
#         if len(missing_list)> 0:
#             for i in missing_list:
#                 missing_sign_message = "Missing "+ i + " Signature on " + date
#                 message_list.append(missing_sign_message)


#                 findings.append(missing_sign_message)
#         else:
#             print("all_signatures_present")

#         if len(message_list)> 0:
#             signatures = ' '.join(message_list)
#             print("Signature message:",message_list)
#             findings.append(message_list)
#             signatures = add_char_every_certain_num_char_in_string(36,signatures)
#             date_note_list.append(signatures)

#         print("appending to excel")
#         append_to_excel(date_note_list)
#         append_count += 1
#         print("APPEND NUM:",append_count)
#         date_note_list = []
#         back(1)  #have to click back to go back once.
#         print("PRESSED BACK B")

#         idg_links = find_elements(idg_links_xps)
#         print("found new idg links")

#     print("GATHERED DATA FROM ALL IDG, Spank you come again")

#     #####################################################################################

# # idg_notes()
# print("IDG NOTES ALL DONE")
# ########################################## Going to VISIT NOTES GETS HTML FOR PARSING###########################################

# def visit_note_collect_html():
#     print("Starting visit_note_collect")
#     click_visit_note_tab()

#     #make sure first row is nursing
#     nurse_abb_list = ['RN',"LVN",'BSN']
#     def find_first_nurse_note():
#         print("******** STARTING find_first_nurse_note**************")
#         visit_note_rows = find_elements('/html/body/form/table/tbody/tr[1]/td/div/table/tbody/tr[2]/td/table/tbody/tr[4]/td[2]/div/table/tbody/tr/td[2]/div[2]/div/table/tbody/tr[*]')
#         print("LEN VISIT NOTE ROWS A: ",len(visit_note_rows)) # DON"T USE FIRST AND LAST

#         x_num =2

#         for row in visit_note_rows[1:-1]:
#             row_text = row.text
#             print(row_text)

#             visit_nurse_row = f'/html/body/form/table/tbody/tr[1]/td/div/table/tbody/tr[2]/td/table/tbody/tr[4]/td[2]/div/table/tbody/tr/td[2]/div[2]/div/table/tbody/tr[{x_num}]/td[1]/input'
#             if "RN" in row_text or "LVN" in row_text or "BSN" in row_text:
#                 click(visit_nurse_row)
#                 print("NUM2:",x_num)
#                 print("clicked first nurse note AA")
#                 break
#             else:
#                 x_num+=1
#                 print("NUM:",x_num)

#     find_first_nurse_note()


#     # nurse_note = click('//*[@id="ctl00_ContentPlaceHolder1_GridView1"]/tbody/tr[3]/td[1]/input', 60)
#     print('Expanding')
#     expand = click('//*[@id="ctl00_ContentPlaceHolder1_ImgExpandAll"]')
#     vital_magnifier = click('//*[@id="ctl00_ContentPlaceHolder1_ImgHistory"]')
#     #this will get row plus another element i don't need so i need to skip every other
#     date = get_value('//*[@id="ctl00_ContentPlaceHolder1_txtHDate"]')
#     print("---------date:------------",date)

#     pt_name = find('//*[@id="ctl00_PB_HEADERPATIENTHEADER"]/b[1]')
#     pt_name =pt_name.text
#     print("pt_name:",pt_name)


#     #added the first vs to load then 2 sleep.. if none it will make error
#     print("sleeping 20 to get page source") # required with slow internet.. fast go to 5
#     sleep(20)

#     visit_note = driver.page_source

#     #saving page to parse it  #cleaning up file name so it saves
#     soup = BeautifulSoup(visit_note, 'html.parser')
#     s = 'visit_note' + date + pt_name
#     s = s.replace('/','.')
#     s = s.replace(',','')
#     comp_filename = s.replace(" ",'')

#     with open(comp_filename, 'wt', encoding='utf-8') as html_file:  #writes code name
#         for line in soup.prettify():
#             html_file.write(line)

#     print("file name: ",comp_filename)
#     return comp_filename

# comp_filename = visit_note_collect_html()


#         ################## GATHERS ALL VITAL SIGNS THAT WERE RECORDED

# def gather_all_vitals(name_of_file):
#     print("CHECKING GATHER_NURSE_NOTE_DATA_BS4")
#     label_list = []
#     mac_list = []
#     oxy_list = []
#     ###################### GATHERING ALL VS FROM first visit Note ###################################
#     with open(name_of_file, 'r') as file:
#             soup = BeautifulSoup(file, 'html.parser')
#             labels = soup.find_all(colspan="9")
#             # print(len(labels))
#             vitals = soup.find_all(id=re.compile("ctl00_ContentPlaceHolder1_MyRepeater_ctl01"))
#             print(len(vitals))

#             vital_num = 1
#             # GETS LABELS
#             for label in labels[1:-1]:  # gonna get 30
#                 label = label.get_text(" ", strip=True)
#                 print("LABEL",label)
#                 label_list.append(label)
#                 date = label[12:21]
#                 nurse = label[24:]


#                 # GETS ALL VITALS ON EACH LINE

#                 if vital_num <= 9:
#                         id_name = f"ctl00_ContentPlaceHolder1_MyRepeater_ctl0{vital_num}"
#                 elif vital_num > 9:
#                         id_name = f"ctl00_ContentPlaceHolder1_MyRepeater_ctl{vital_num}"
#                 # print("ID NAME: ",id_name)
#                 # print(id_name)
#                                 # ctl00_ContentPlaceHolder1_MyRepeater_ctl01_txt
#                 vitals = soup.find_all(id=re.compile(id_name))
#                 vital_num +=1
#                 print("VITAL LEN:",len(vitals))
#                 # print("vital num1:",vital_num)
#                 findvs= 0                        # CURRENT PROBLEM DOING ALL TEMP NOT GOING To NEXT VS
#                 visit_list = []
#                 visit_list.append(date)
#                 visit_list.append(nurse)

#                 for vital in vitals:
#                         findvs += 1

#                         try:  # COULDN"T FIND ALL TOGETHER IN BS$ SO FILTERING HEREE
#                             # if i want more data just print the findvs num  plus value and you'll see all options
#                             vs = vital['value']
#                             # print('Good:',good,vs)
#                             found = ''
#                             if findvs == 3:
#                                 found = 'yes'
#                                 vital_name = "temp"
#                                 print(vital_name,vs)
#                             if findvs == 4:
#                                 found = 'yes'
#                                 vital_name = "pulse"
#                                 print("Pulse:",vs)
#                             if findvs == 5:
#                                 found = 'yes'
#                                 vital_name = "r"
#                                 print("R:",vs)
#                             if findvs == 6:
#                                 found = 'yes'
#                                 vital_name = "bp"
#                                 print("BP:",vs)
#                                 result = check_bp_range(vs)

#                                 if len(result) > 0:
#                                     print("Found BP out of range")
#                                     findings.append(result)

#                             if findvs == 13:
#                                 found = 'yes'
#                                 vital_name = "weight"
#                                 print('Weight',vs)
#                             if findvs == 14:
#                                 found = 'yes'
#                                 vital_name = "mac"
#                                 print("Mac:",vs)
#                                 mac_list.append(float(vs))

#                             if findvs == 15:
#                                 found = 'yes'
#                                 vital_name = "Arm"
#                                 print("Arm:",vs)
#                             if findvs == 20:
#                                 found = 'yes'
#                                 vital_name = "02 sat"
#                                 print('02 Sat:',vs)
#                                 oxy_list.append(float(vs))
#                             # adds name and value to visit-vitals dict
#                             if found == 'yes':
#                                 # print("vital_name",vital_name)
#                                 print("vs",vs)

#                                 visit_list.append(vs)

#                         except:
#                             # print("bad num so passed")
#                             pass

#                 ws.append(visit_list)
#                 wb.save(excel_location)
#             print("ALL VITAL SIGNS APPENDED TO EXCEL:",excel_location)

#     return mac_list,oxy_list
# ################################################ END OF GATHERING VISIT NOTE VITALS################################
# returned = gather_all_vitals(comp_filename)
# # print("RETURNED: ",returned)
# mac_list = returned[0]
# # print("mac_list: ",mac_list)
# oxy_list = returned[1]
# # print("oxy_list: ",oxy_list)
# from statistics import mean

# try:
#     avg_oxy = mean(oxy_list)
# except:
#     print('avg-oxy list blank')

# try:
#     lowest_mac = min(mac_list)
# except:
#     print("lowest_mac list empty")
#     #Gather all data from visit note through bs4 except all vitals
# def gather_visit_note_details():
#     print("Starting gather visit notes")
#     ###################### GATHERING ALL VS FROM first visit Note file ###################################

#     ########## STARTING TO ITERATE TABS OF VISIT NOTES ################
#     click_visit_note_tab()
#     page_num = 1
#     ############ NEED TO SPECIFY EACH ELEMENT IN FIND FUNCTION FOR FOLLOWING ELEMENTS
#     #required because it saves the number that needs to iterate to the correct link
#     page_num_links = '/html/body/form/table/tbody/tr[1]/td/div/table/tbody/tr[2]/td/table/tbody/tr[4]/td[2]/div/table/tbody/tr/td[2]/div[2]/div/table/tbody/tr[17]/td/table/tbody/tr/td[*]'

#     multiple_pages = ''

#     try:
#         pages = find_elements(page_num_links)
#         number_of_pages = len(pages)
#         multiple_pages = 'yes'
#         print("LEN PAGES: ",number_of_pages)
#     except:
#         multiple_pages = 'no'

#     #clicks all pages one at a time
#     if multiple_pages == 'yes':
#         print("yes multiple pages")

#         for i in range(len(pages)):

#             print("Page Loop a:",i)
#             print("Sleep 5...")
#             sleep(5)
#             print("page_num",page_num)
#             #CLICKS PAGE THEN CLICKS ALL good visit links IN THEM
#             click(f'/html/body/form/table/tbody/tr[1]/td/div/table/tbody/tr[2]/td/table/tbody/tr[4]/td[2]/div/table/tbody/tr/td[2]/div[2]/div/table/tbody/tr[17]/td/table/tbody/tr/td[{page_num}]')
#             page_num +=1
#             print("clicked next page")

#             link_counter = 2 #starting 2 since xpath starts w 2
#             date_note_list = []
#             x_num_list = []   #gotta clean list each iteration
#             x_num =2   #
#         #############ORIGINAL CODE ##############

#             ##### GATHERS ROW NUMBERS to LIST TO CREATE XPATHS BELOW

#             visit_note_rows = find_elements('/html/body/form/table/tbody/tr[1]/td/div/table/tbody/tr[2]/td/table/tbody/tr[4]/td[2]/div/table/tbody/tr/td[2]/div[2]/div/table/tbody/tr[*]')
#             # page_url was heres

#             print("LEN VISIT NOTE ROWS: ",len(visit_note_rows))
#                         ###### SLICED BECAUSE CAN'T USE FIRST AND LAST
#             for row in visit_note_rows[1:-1]:
#                 row_text = row.text
#                 # print(row_text)

#                 if "RN" in row_text or "LVN" in row_text or "BSN" in row_text:

#                     x_num_list.append(x_num)
#                     x_num+= 1
#                     # print("X_NUM:",x_num)
#                 else:
#                     x_num+= 1
#                     print("no num")

#                     ##NOW CHANGE XPATH FROM THE LIST ITERATION. AND GATHER FROM PAGES AND ALL

#             #CLICKS ON VISIT NOTE TO GATHER DATA

#             for num in x_num_list:  #
#                 print("Starting to click visit_nurse row A :",num)
#                 visit_note_list = []  #should collect everything then refresh it
#                 visit_nurse_row = f'/html/body/form/table/tbody/tr[1]/td/div/table/tbody/tr[2]/td/table/tbody/tr[4]/td[2]/div/table/tbody/tr/td[2]/div[2]/div/table/tbody/tr[{num}]/td[1]/input'
#                 # print("VISIT NURSE ROW: ",visit_nurse_row)

#                 click(visit_nurse_row)
#                 print("SHOUDLVE clicked it A")
#                 print("NUM2:",x_num)
#                 # expand = click('//*[@id="ctl00_ContentPlaceHolder1_ImgExpandAll"]')

#                 # print("Trying new load technique so i don't have to use waits")

#                     #using this instead of sleep
#                 url1 = driver.current_url
#                 new = url1[:-3]+ 'c=&drl=exp'
#                 print("URL NEW",new)
#                 driver.get(new)

#                 visit_note2 = driver.page_source
#                 soup = BeautifulSoup(visit_note2, 'html.parser')

#                 comp_filename2 = "visit note"

#                 with open(comp_filename2, 'wt', encoding='utf-8') as html_file:  #writes code name
#                     for line in soup.prettify():
#                             html_file.write(line)

#                 print("file Visit Note Sourcepage saved as: ",comp_filename2)  # ADDED THE 2 on this

#                 with open(comp_filename2, 'r',encoding='utf-8') as file:
#                     soup = BeautifulSoup(file, 'html.parser')

#                     def get_data(id):
#                         try:
#                             item = soup.find(id=re.compile(id))
#                             if 'type="checkbox"' in str(item):
#                                     if 'checked="checked"' in str(item):
#                                         return 'checked'
#                                     else:
#                                         return "not checked"

#                             selected = item.select_one('option:checked')
#                             value = selected.get_text("|", strip=True)
#                         except:
#                             try:
#                                     item = soup.find(id=re.compile(id))
#                                     value = item['value']

#                             except:
#                                     item = soup.find(id=re.compile(id))
#                                     value = item.get_text("|", strip=True)

#                         if value == '1' or value == '2':
#                             if 'checked="checked"' in str(item):
#                                     print('checked')
#                                     return 'checked'

#                         visit_note_list.append(value)
#                         # print(value)
#                         return value

#                     def get_radio(id,print_name):
#                         item = soup.find(id=re.compile(id))
#                         if 'checked="checked"' in str(item):
#                             name = print_name
#                             visit_note_list.append(name)
#                             print(name)
#                             value = 'checked'
#                         else:
#                             value = ''
#                         return value

#                     def get_text_value(id):
#                         item = soup.find(id=re.compile(id))
#                         value = item['value']
#                         # print("Value")
#                         return value


#                     def collecting():  #added this just so i can keep it all together
#                         print("----------------collecting------------------")
#                         visit_date = get_text_value('ctl00_ContentPlaceHolder1_txtHDate')
#                         visit_note_list.append(visit_date)

#                         print("VISIT_DATE:",visit_date)

#                         nurse = get_data('ctl00_ContentPlaceHolder1_DropCreatedBy')
#                         # nurse = get_data('ctl00_ContentPlaceHolder1_DropCreatedBy')
#                         discipline = get_data('ctl00_ContentPlaceHolder1_DropDiscipline')
#                         care_level = get_data('ctl00_ContentPlaceHolder1_DropCareLevel')

#                         narrative = soup.find(id=re.compile('ctl00_ContentPlaceHolder1_TxtComments'))
#                         narrative = narrative.get_text("|", strip=True)
#                         narrative = add_char_every_certain_num_char_in_string(50,narrative)

#                         if len(narrative) < 5:
#                             findings.append("Missing Nurse Narrative by "+ nurse + ' '+visit_date)


#                         visit_note_list.append(narrative)
#                         teaching_list = ['safety','skin','fall']
#                         for word in teaching_list:
#                             if word in narrative:
#                                 pass
#                             else:
#                                 print("Nurse did not put '"+ word +"'in narrative.") #append to findings
#                                 findings.append("Nurse did not put '"+ word +"'in narrative."+ visit_date)
#                                 break

#                         # print(narrative)
#                         # narrative = narrative.get_text("|", strip=True)
#                         # print(narrative)

#                         form_type = get_data('ctl00_ContentPlaceHolder1_DrpVisitUnsc')
#                         del visit_note_list[-1]
#                         visit_note_list.append('Form Type:'+form_type)

#                         # frequency = get_data('ctl00_ContentPlaceHolder1_lblFrequencyOfVisit')
#                         chief_complaint = get_data('ctl00_ContentPlaceHolder1_TxtPsgStatesChiefComplaint')

#                         ############## COVID #########################
#                         covid_screen_yes = get_data('ctl00_ContentPlaceHolder1_rbPriorScreening_1')

#                         # del visit_note_list[-1]
#                         if covid_screen_yes == 'checked':
#                             pass

#                         else:
#                             findings.append('Covid Screen not done. '+ visit_date)
#                             print("covid screen not done.. ")
#                             click_note('//*[@id="ctl00_ContentPlaceHolder1_rbPriorScreening_1"]')


#                         test_date = get_data('ctl00_ContentPlaceHolder1_txtPriorScreenDate')
#                         # del visit_note_list[-1]
#                         if test_date == '':
#                             findings.append("Covid test date not entered")
#                             type_note('ctl00$ContentPlaceHolder1$txtPriorScreenDate',visit_date)

#                         pt_report_positive = get_data('ctl00_ContentPlaceHolder1_rbPositive_0')
#                         # del visit_note_list[-1]
#                         if pt_report_positive != 'checked': findings.append("pt reported positive"+ visit_date)

#                         pt_report_fever = get_data('ctl00_ContentPlaceHolder1_rbCoronaVirus_0')
#                         # del visit_note_list[-1]
#                         if pt_report_positive != 'checked': findings.append("pt report fever"+ visit_date)

#                         add_comment = get_data('ctl00_ContentPlaceHolder1_rbPatientPcgVisit_0')
#                         # del visit_note_list[-1]
#                         if pt_report_positive != 'checked': findings.append("additional comment in covid"+ visit_date)

#                         followed_covid = get_data('ctl00_ContentPlaceHolder1_chkCoronaInstructions')
#                         del visit_note_list[-1]
#                         if followed_covid != 'checked':
#                             findings.append("pt reported covid positive"+ visit_date)
#                             print("pt reported positive not checked")
#                             click_note('//*[@id="ctl00_ContentPlaceHolder1_chkCoronaInstructions"]')


#                         utilized_ppe = get_data('ctl00_ContentPlaceHolder1_chkPpe')
#                         del visit_note_list[-1]
#                         if utilized_ppe != 'checked':
#                             findings.append("nurse didn't select box utilized ppe"+ visit_date)
#                             print("utilized ppe not checked")
#                             click_note('//*[@id="ctl00_ContentPlaceHolder1_chkPpe"]')


#                         temp = get_value('//*[@id="ctl00_ContentPlaceHolder1_txtTemp"]')

#                         if len(temp) < 1:
#                             print(temp)
#                             findings.append("No Temp"+ visit_date)
#                             type_note('//*[@id="ctl00_ContentPlaceHolder1_txtTemp"]','97.6')


#                         pulse = get_value('//*[@id="ctl00_ContentPlaceHolder1_TxtPulse"]')
#                         if len(pulse) < 1:
#                             print(pulse)
#                             findings.append('No Pulse' + visit_date)
#                             type_note('//*[@id="ctl00_ContentPlaceHolder1_TxtPulse"]','72')
#                             # typing('//*[@id="ctl00_ContentPlaceHolder1_TxtPulse"]',3,'72')

#                         resp = get_value('//*[@id="ctl00_ContentPlaceHolder1_TxtResp"]')
#                         if len(resp) < 1:
#                             findings.append("No Respirations"+ visit_date)
#                             type_note('//*[@id="ctl00_ContentPlaceHolder1_TxtResp"]','17')
#                             # typing('//*[@id="ctl00_ContentPlaceHolder1_TxtResp"]',3,'18')

#                         mac1 = get_value('//*[@id="ctl00_ContentPlaceHolder1_TxtMac"]')
#                         if len(mac1) < 1:
#                             findings.append("No Mac",visit_date)
#                             type_note('//*[@id="ctl00_ContentPlaceHolder1_TxtMac"]',lowest_mac)
#                             # typing('//*[@id="ctl00_ContentPlaceHolder1_TxtMac"]',3,lowest_mac)
#                             print("Typing no mac")

#                         oxy = get_value('//*[@id="ctl00_ContentPlaceHolder1_TxtO2SAT"]')
#                         if len(oxy) < 1:
#                             findings.append("No oxygen sat"+ visit_date)
#                             print("Nooxygen sat")
#                             type_note('//*[@id="ctl00_ContentPlaceHolder1_TxtO2SAT"]',avg_oxy)
#                             # typing('//*[@id="ctl00_ContentPlaceHolder1_TxtO2SAT"]',3,avg_oxy)


#                         ################################# PPE######################################

#                         is_pain_controlled = get_data('ctl00_ContentPlaceHolder1_RbControlled_0')

#                         del visit_note_list[-1]
#                         if is_pain_controlled != 'Yes': findings.append('Pain NOT controlled')

#                         pain_reported = get_data('ctl00_ContentPlaceHolder1_DropPainLevel')
#                         del visit_note_list[-1]
#                         visit_note_list.append("Pain: "+pain_reported)


#                         ##################### SIGNS AND SYMPTOMS #############################

#                         ################ NEURO ###############
#                         #PRINTS OUT THE ACTUAL SELECTED ONE

#                         def if_none_checked_add_findings(category,list):
#                             num = 0
#                             for var in list:
#                                 if var == 'checked':
#                                         num += 1
#                                         pass
#                                 else:
#                                         pass
#                             if num == 0:
#                                 findings.append(category + ' boxes were not selected in visit notes')

#                         #STILL NEED TO ADD TEXT INPUTS TO THESE
#                         if form_type != 'Short Form':
#                             print("NOT SHORT FORM**************")
#                             anxiety = get_radio('ctl00_ContentPlaceHolder1_RB_NMS_Anx1','anxiety none')
#                             anxiety_mild = get_radio('ctl00_ContentPlaceHolder1_RB_NMS_Anx2','anxiety_mild')
#                             anxiety_mod = get_radio('ctl00_ContentPlaceHolder1_RB_NMS_Anx3','anxiety_mod')
#                             anxiety_severe = get_radio('ctl00_ContentPlaceHolder1_RB_NMS_Anx4','anxiety_severe')
#                             anxiety_list = [anxiety,anxiety_mild,anxiety_mod,anxiety_severe]
#                             if_none_checked_add_findings("Anxiety",anxiety_list)

#                             depression_none= get_radio('ctl00_ContentPlaceHolder1_RB_NMS_Dep1','depression_none')
#                             dep_mild = get_radio('ctl00_ContentPlaceHolder1_RB_NMS_Dep2','dep_mild')
#                             dep_mod = get_radio('ctl00_ContentPlaceHolder1_RB_NMS_Dep3','dep_mod')
#                             dep_severe = get_radio('ctl00_ContentPlaceHolder1_RB_NMS_Dep4','dep_severe')
#                             depression_list = [depression_none,dep_mild,dep_mod,dep_severe]
#                             if_none_checked_add_findings("Depression",depression_list)

#                             agitation_none = get_radio('ctl00_ContentPlaceHolder1_RB_NMS_DIZ1','agitation_none')
#                             agitation_mild = get_radio('ctl00_ContentPlaceHolder1_RB_NMS_DIZ2','agitation_mild')
#                             agitation_mod = get_radio('ctl00_ContentPlaceHolder1_RB_NMS_DIZ3','agitation_mod')
#                             agitation_severe = get_radio('ctl00_ContentPlaceHolder1_RB_NMS_DIZ4','agitation_severe')
#                             agitation_list = [agitation_none,agitation_mild,agitation_mod,agitation_severe]
#                             if_none_checked_add_findings("Agitation",agitation_list)

#                             confusion_none = get_radio('ctl00_ContentPlaceHolder1_RBS_NMS_CON1','confusion_none')
#                             confusion_mild = get_radio('ctl00_ContentPlaceHolder1_RBS_NMS_CON2','confusion_mild')
#                             confusion_mod = get_radio('ctl00_ContentPlaceHolder1_RBS_NMS_CON3','confusion_mod')
#                             confusion_sev = get_radio('ctl00_ContentPlaceHolder1_RBS_NMS_CON4','confusion_sev')
#                             confusion_list = [confusion_none,confusion_mild,confusion_mod,confusion_sev]
#                             if_none_checked_add_findings("confusion",confusion_list)

#                             total_adl_score = 'ctl00_ContentPlaceHolder1_TotalADL'
#                             item = soup.find(id=re.compile(total_adl_score))
#                             total_adl = item.get_text("|", strip=True)
#                             visit_note_list.append("Total ADL Score: "+total_adl)
#                             print("TOTAL_ADL_SCORE",total_adl)

#                             if 'Alzheimer' in diagnosis:
#                                 sev = get_checked('//*[@id="ctl00_ContentPlaceHolder1_RBS_NMS_CON4"]')
#                                 # print("SEVERE: ",sev)
#                                 # print(type(sev))
#                                 if sev != 'true':
#                                     no_issues = get_checked('//*[@id="ctl00_ContentPlaceHolder1_ChkNeuological"]')
#                                     if no_issues == 'true':
#                                         print("PT has Alzheimers and nurse selected no issues in neuro "+visit_date)
#                                         findings.append("PT has Alzheimers and nurse selected no issues in neuro "+visit_date)
#                                         click_note('//*[@id="ctl00_ContentPlaceHolder1_ChkNeuological"]')

#                                     print('nurse did not select pt has severe confusion.. fixing it.')
#                                     findings.append("Pt has alzheimers and nurse didn't select severe confusion"+visit_date)
#                                     click_note('//*[@id="ctl00_ContentPlaceHolder1_RBS_NMS_CON4"]')
#                                     # click('//*[@id="ctl00_ContentPlaceHolder1_RBS_NMS_CON4"]')

#                                 urinary_incontinent = get_selected('//*[@id="ctl00_ContentPlaceHolder1_Gen_URCin"]','urinary_incontinent')
#                                 if urinary_incontinent != 'Yes':
#                                     print("urinary incont no")
#                                     findings.append("Urinary_incont not selected yes with Alzheimers pt "+visit_date)
#                                     click_note('//*[@id="ctl00_ContentPlaceHolder1_Gen_URCin"]/option[2]')
#                                     # yes = click('//*[@id="ctl00_ContentPlaceHolder1_Gen_URCin"]/option[2]')

#                                 bowel_incont = get_selected('//*[@id="ctl00_ContentPlaceHolder1_Gas_STS"]',"bowel incont")
#                                 if bowel_incont != 'Yes':
#                                     print('vowel incon no')
#                                     findings.append("Bowel_incont not selected yes with Alzheimers pt "+visit_date)
#                                     click_note('//*[@id="ctl00_ContentPlaceHolder1_Gas_STS"]')
#                                     # yes = click('//*[@id="ctl00_ContentPlaceHolder1_Gas_STS"]')

#                                 ambulation = get_selected('//*[@id="ctl00_ContentPlaceHolder1_DRP_CEAD_A"]','ambulation')
#                                 if ambulation != '3':
#                                     print("ambul not 3")
#                                     findings.append("ambulation wasn't 3 on "+visit_date)
#                                     ambu = click_note('//*[@id="ctl00_ContentPlaceHolder1_DRP_CEAD_A"]/option[4]')

#                                 dressing = get_selected('//*[@id="ctl00_ContentPlaceHolder1_DRP_CEAD_D"]','dressing')
#                                 if dressing != '3':
#                                     print('dress not 3')
#                                     findings.append("dressing wasn't 3 on "+visit_date)
#                                     dress = click_note('//*[@id="ctl00_ContentPlaceHolder1_DRP_CEAD_D"]/option[4]')


#                                 bathing = get_selected('//*[@id="ctl00_ContentPlaceHolder1_DRP_CEAD_B"]','bathing')
#                                 if bathing != '3':
#                                     print('bath  not 3')
#                                     findings.append("bathing wasn't 3 on "+visit_date)
#                                     bath = click_note('//*[@id="ctl00_ContentPlaceHolder1_DRP_CEAD_B"]/option[4]')

#                                     #WHERE IS FAST"?
#                                 # fast = get_selected('//*[@id="ctl00_ContentPlaceHolder1_DDFast"]','fast')
#                                 # fast_list = ['7-A','7-B','7-C','7-D','7-E','7-F']
#                                 # if fast in fast_list:
#                                 #     pass
#                                 # else:
#                                 #     print("FAST SCORE NOT APPROPRIATE FOR ALZHEIMER PT",fast,visit_date)
#                                 #     findings.append("Fast score "+ fast +" for Alzheimer pt " + visit_date)
#                                 #     click_note('//*[@id="ctl00_ContentPlaceHolder1_DDFast"]/option[12]')
#                                 #     print("shouldve changed it.")
#                                 #########WORK ON THIS


#                             speech_impair_none = get_radio('ctl00_ContentPlaceHolder1_RBS_SI_CON1','speech_impair_none')
#                             speech_impair_mild = get_radio('ctl00_ContentPlaceHolder1_RBS_SI_CON2','speech_impair_mild')
#                             speech_impair_mod = get_radio('ctl00_ContentPlaceHolder1_RBS_SI_CON3','speech_impair_mod')
#                             speech_impair_sev = get_radio('ctl00_ContentPlaceHolder1_RBS_SI_CON4','speech_impair_sev')
#                             speech_list = [speech_impair_none,speech_impair_mild,speech_impair_mod,speech_impair_sev]
#                             if_none_checked_add_findings("Speech",speech_list)

#                             other_symp_none = get_radio('ctl00_ContentPlaceHolder1_RB_NMS_OSY1','other_sym_none')
#                             other_mild = get_radio('ctl00_ContentPlaceHolder1_RB_NMS_OSY2','other_sym_mild')
#                             other_mod = get_radio('ctl00_ContentPlaceHolder1_RB_NMS_OSY3','other_symp_mod')
#                             other_sev = get_radio('ctl00_ContentPlaceHolder1_RB_NMS_OSY4','other_symp_sev')
#                             other_list = [other_symp_none,other_mild,other_mod,other_sev]
#                             if_none_checked_add_findings('Other Symptoms',other_list)

#                             #################### CARDIO#################################

#                             arrhythmia_none = get_radio('ctl00_ContentPlaceHolder1_RB_CAR_ARR1', "arrhythmia_none")
#                             arr_mild = get_radio("ctl00_ContentPlaceHolder1_RB_CAR_ARR2","Arrythmia Mild")
#                             arr_mod = get_radio('ctl00_ContentPlaceHolder1_RB_CAR_ARR3','Arrythmia mod')
#                             arr_sev = get_radio('ctl00_ContentPlaceHolder1_RB_CAR_ARR4',"Arrhythmia Severe")
#                             arrhythmia_list = [arrhythmia_none,arr_mild,arr_mod,arr_sev]
#                             if_none_checked_add_findings("Arrhythmia", arrhythmia_list)

#                             edema_none = get_radio('ctl00_ContentPlaceHolder1_RB_CAR_EDM1','Edema None')
#                             edema_mild = get_radio('ctl00_ContentPlaceHolder1_RB_CAR_EDM2',"Edema Mild")
#                             edema_mod = get_radio('ctl00_ContentPlaceHolder1_RB_CAR_EDM3','Edema Mod')
#                             edema_sev = get_radio('ctl00_ContentPlaceHolder1_RB_CAR_EDM4',"Edema Severe")
#                             edema_list = [edema_none,edema_mild,edema_mod,edema_sev]
#                             if_none_checked_add_findings("EDEMA", edema_list)

#                             chest_pain_none = get_radio('ctl00_ContentPlaceHolder1_RB_CAR_CHP1',"No Chest Pain")
#                             chest_pain_mild = get_radio('ctl00_ContentPlaceHolder1_RB_CAR_CHP2',"Mild chest pain")
#                             chest_pain_mod = get_radio('ctl00_ContentPlaceHolder1_RB_CAR_CHP3',"Moderate chest pain")
#                             chest_pain_sev = get_radio('ctl00_ContentPlaceHolder1_RB_CAR_CHP4',"severe chest pain")
#                             chest_pain_oth = get_radio('ctl00_ContentPlaceHolder1_RB_CAR_OSY1',"chest pain other")
#                             chest_pain_list = [chest_pain_none,chest_pain_mild,chest_pain_mod,chest_pain_sev,chest_pain_oth]
#                             if_none_checked_add_findings("Chest Pain",chest_pain_list)

#                             ################# RESPIRATORY #######################
#                             dyspnea_none = get_radio('ctl00_ContentPlaceHolder1_RB_CAR_DYS1',"Dyspnea None")
#                             dyspnea_mild = get_radio('ctl00_ContentPlaceHolder1_RB_CAR_DYS2',"Dyspnea Mild")
#                             dyspnea_mod = get_radio('ctl00_ContentPlaceHolder1_RB_CAR_DYS3',"Dyspnea Mod")
#                             dyspnea_sev = get_radio('ctl00_ContentPlaceHolder1_RB_CAR_DYS4',"Dyspnea Sev")
#                             dyspnea_list = [dyspnea_none,dyspnea_mild,dyspnea_mod,dyspnea_sev]
#                             if_none_checked_add_findings("Dyspnea",dyspnea_list)

#                             cough_none = get_radio('ctl00_ContentPlaceHolder1_RB_C1',"Cough None")
#                             cough_mild = get_radio('ctl00_ContentPlaceHolder1_RB_C2',"Cough Mild")
#                             cough_mod = get_radio('ctl00_ContentPlaceHolder1_RB_C3',"Cough Mod")
#                             cough_sev = get_radio('ctl00_ContentPlaceHolder1_RB_C4',"Cough Severe")
#                             cough_list = [cough_none,cough_mild,cough_mod,cough_sev]
#                             if_none_checked_add_findings('Cough',cough_list)

#                             abnormal_lung_sounds = get_radio('ctl00_ContentPlaceHolder1_RB_A1', "Abnormal Lung Sounds None")
#                             lung_sounds_mild = get_radio('ctl00_ContentPlaceHolder1_RB_A2',"Abnormal Lung Sounds Mild")
#                             lung_sounds_mod = get_radio('ctl00_ContentPlaceHolder1_RB_A3',"Abnormal Lung Sounds Mild")
#                             lung_sounds_sev = get_radio('ctl00_ContentPlaceHolder1_RB_A4',"Abnormal lung sounds Severe")
#                             lung_sound_list = [abnormal_lung_sounds,lung_sounds_mild,lung_sounds_mod,lung_sounds_sev]
#                             if_none_checked_add_findings('Lung Sounds',lung_sound_list)

#                             lung_sound_other_none = get_radio('ctl00_ContentPlaceHolder1_RB_RES_OSY1',"Lung Sounds Other NONE")
#                             ls_other_mild = get_radio('ctl00_ContentPlaceHolder1_RB_RES_OSY2', "Lung Sounds other MILD")
#                             ls_other_mod = get_radio('ctl00_ContentPlaceHolder1_RB_RES_OSY3',"Lung Sounds other MOD")
#                             ls_other_sev = get_radio('ctl00_ContentPlaceHolder1_RB_RES_OSY4',"Lung Sounds other SEVERE")
#                             ls_other_list = [lung_sound_other_none,ls_other_mild,ls_other_mod,ls_other_sev]
#                             if_none_checked_add_findings("Lung Sounds other",ls_other_list)

#                             ################ IMMUNOLOGICAL ################################
#                             infection_none = get_radio('ctl00_ContentPlaceHolder1_RB_IMG_INF1',"INFECTION NONE")
#                             infection_mild = get_radio('ctl00_ContentPlaceHolder1_RB_IMG_INF2',"Infection MILD")
#                             infection_mod = get_radio('ctl00_ContentPlaceHolder1_RB_IMG_INF3',"Infection MOD")
#                             infection_sev = get_radio('ctl00_ContentPlaceHolder1_RB_IMG_INF4',"Infection SEVERE")
#                             infection_list = [infection_none,infection_mild,infection_mod,infection_sev]
#                             if_none_checked_add_findings('Infections',infection_list)

#                             ############### GASTRO INTESTINAL ###############################

#                             nausea_none = get_radio('ctl00_ContentPlaceHolder1_RB_GAS_NAU1',"Nausea None")
#                             nausea_mild = get_radio('ctl00_ContentPlaceHolder1_RB_GAS_NAU2',"Nausea MILD")
#                             nausea_mod = get_radio('ctl00_ContentPlaceHolder1_RB_GAS_NAU3',"Nausea MOD")
#                             nausea_sev = get_radio('ctl00_ContentPlaceHolder1_RB_GAS_NAU4',"Nausea SEV")
#                             nausea_list = [nausea_none,nausea_mild,nausea_mod,nausea_sev]
#                             if_none_checked_add_findings("Nausea",nausea_list)

#                             vomiting_none = get_radio('ctl00_ContentPlaceHolder1_RB_GAS_VOM1',"vomiting_none")
#                             vomit_mild = get_radio('ctl00_ContentPlaceHolder1_RB_GAS_VOM2',"vomit_mild")
#                             vomit_mod = get_radio('ctl00_ContentPlaceHolder1_RB_GAS_VOM3',"vomit_mod")
#                             vomit_sev = get_radio('ctl00_ContentPlaceHolder1_RB_GAS_VOM4',"vomit_sev")
#                             vomit_list = [vomiting_none,vomit_mild,vomit_mod,vomit_sev]
#                             if_none_checked_add_findings("Vomitting", vomit_list)

#                             constipation_none = get_radio('ctl00_ContentPlaceHolder1_RB_GAS_CON1', "Constipation NONE")
#                             con_mild = get_radio('ctl00_ContentPlaceHolder1_RB_GAS_CON2','Constipation MILD')
#                             con_mod = get_radio('ctl00_ContentPlaceHolder1_RB_GAS_CON3',"Constipation MOD")
#                             con_sev = get_radio('ctl00_ContentPlaceHolder1_RB_GAS_CON4',"Constipation SEVERE")
#                             constipation_list = [constipation_none,con_mild,con_mod,con_sev]
#                             if_none_checked_add_findings("Constipation",constipation_list)

#                             diarrhea_none = get_radio('ctl00_ContentPlaceHolder1_RB_GAS_DIA1',"Diarrhea NONE")
#                             diarrhea_mild = get_radio('ctl00_ContentPlaceHolder1_RB_GAS_DIA2',"Diarrhea MILD")
#                             diarrhea_mod = get_radio('ctl00_ContentPlaceHolder1_RB_GAS_DIA3',"Diarrhea MODERATE")
#                             diarrhea_sev = get_radio('ctl00_ContentPlaceHolder1_RB_GAS_DIA4',"Diarrhea SEVERE")
#                             diarrhea_list = [diarrhea_none,diarrhea_mild,diarrhea_mod,diarrhea_sev]
#                             if_none_checked_add_findings("Diarrhea",diarrhea_list)

#                             other_gi_none = get_radio('ctl00_ContentPlaceHolder1_RB_GAS_DIA1',"Other GI NONE")
#                             other_gi_mild = get_radio('ctl00_ContentPlaceHolder1_RB_GAS_DIA2',"Other GI Mild")
#                             other_gi_mod = get_radio('ctl00_ContentPlaceHolder1_RB_GAS_DIA3',"other GI Mod")
#                             other_gi_sev = get_radio('ctl00_ContentPlaceHolder1_RB_GAS_DIA4',"Other GI SEVERE")
#                             other_gi_list = [other_gi_none,other_gi_mild,other_gi_mod,other_gi_sev]
#                             if_none_checked_add_findings("Other GI",other_gi_list)

#                             last_bm = get_data("ctl00_ContentPlaceHolder1_LastBM")
#                             bm_last = "Last BM: " + last_bm

#                             del visit_note_list[-1]
#                             visit_note_list.append(bm_last)

#                             incontinent_id = 'ctl00_ContentPlaceHolder1_Gas_STS'
#                             incontinent = get_data(incontinent_id)
#                             incontinent_sent = 'Incontinent: '+ incontinent
#                             visit_note_list.append(incontinent_sent)


#                             ############## NUTRITION #################

#                             perc_intake_none = get_radio("ctl00_ContentPlaceHolder1_RB_GAS_OI1", "Perc Oral Intake NONE")
#                             perc_intake_mild = get_radio('ctl00_ContentPlaceHolder1_RB_GAS_OI2',"Perc Oral Intake Mild")
#                             perc_intake_mod = get_radio("ctl00_ContentPlaceHolder1_RB_GAS_OI3","perc Oral Intake Mod")
#                             perc_intake_sev = get_radio('ctl00_ContentPlaceHolder1_RB_GAS_OI4',"Perc intake SEVERE")
#                             intake_list = [perc_intake_none,perc_intake_mild,perc_intake_mod,perc_intake_sev]
#                             if_none_checked_add_findings("Percentage of Oral Intake",intake_list)

#                             nutrition_other_none = get_radio("ctl00_ContentPlaceHolder1_RB_NUT_OSY1","Nutrition Other NONE")
#                             nut_other_mild = get_radio('ctl00_ContentPlaceHolder1_RB_NUT_OSY2',"Nutrition Other MILD")
#                             nut_other_mod = get_radio('ctl00_ContentPlaceHolder1_RB_NUT_OSY3',"Nutrition Other MOD")
#                             nut_other_sev = get_radio('ctl00_ContentPlaceHolder1_RB_NUT_OSY4',"Nutrition other SEVERE")
#                             nut_other_list = [nutrition_other_none,nut_other_mild,nut_other_mod,nut_other_sev]
#                             if_none_checked_add_findings("Nutrition Other",nut_other_list)

#                             artificially_fed_no = get_radio("ctl00_ContentPlaceHolder1_Gas_AFF_0","Artifically fed No")
#                             fed_peg = get_radio('ctl00_ContentPlaceHolder1_Gas_AFF_1',"PEG Feeding")
#                             fed_ng = get_radio('ctl00_ContentPlaceHolder1_Gas_AFF_2',"NG Tube")
#                             fed_j_tube = get_radio('ctl00_ContentPlaceHolder1_Gas_AFF_3',"J-Tube")
#                             fed_pump = get_radio("ctl00_ContentPlaceHolder1_Gas_AFF_4","Pump Feeding")
#                             fed_tpn = get_radio('ctl00_ContentPlaceHolder1_Gas_AFF_5',"TPN Feeding")
#                             art_fed_list = [artificially_fed_no,fed_peg,fed_j_tube,fed_pump,fed_tpn]
#                             if_none_checked_add_findings("Artificially fed",art_fed_list)
#                             diet = get_data("ctl00_ContentPlaceHolder1_DrpDietType")

#                             ########### ENDOCRINE #####################
#                             bs_none = get_radio('ctl00_ContentPlaceHolder1_RB_IND_DIA1',"Blood Sugar NONE")
#                             bs_mild = get_radio('ctl00_ContentPlaceHolder1_RB_IND_DIA2',"Mild Blood Sugar")
#                             bs_mod = get_radio('ctl00_ContentPlaceHolder1_RB_IND_DIA3',"Mod Blood Sugar")
#                             bs_sev = get_radio("ctl00_ContentPlaceHolder1_RB_IND_DIA4","Severe BS")
#                             bs_list = [bs_none,bs_mild,bs_mod,bs_sev]
#                             if_none_checked_add_findings("Blood Sugar",bs_list)

#                             other_bs_none = get_radio('ctl00_ContentPlaceHolder1_RB_IND_OSY1',"other BS NONE")
#                             other_bs_mild = get_radio('ctl00_ContentPlaceHolder1_RB_IND_OSY2',"other BS MILD")
#                             other_bs_mod = get_radio('ctl00_ContentPlaceHolder1_RB_IND_OSY3',"Other BS MOD")
#                             other_bs_sev = get_radio('ctl00_ContentPlaceHolder1_RB_IND_OSY4',"Other BS SEVERE")
#                             other_bs_list = [other_bs_none,other_bs_mild,other_bs_mod,other_bs_sev]
#                             if_none_checked_add_findings("Other BS",other_bs_list)

#                             ####### GENITOURINARY REPRODUCTIVE ###########

#                             urine_none = get_radio('ctl00_ContentPlaceHolder1_RB_GEN_URP1',"Urine No issues")
#                             urine_mild = get_radio('ctl00_ContentPlaceHolder1_RB_GEN_URP2',"Urine MILD")
#                             urine_mod = get_radio('ctl00_ContentPlaceHolder1_RB_GEN_URP3',"Urine MODERATE")
#                             urine_sev = get_radio('ctl00_ContentPlaceHolder1_RB_GEN_URP4',"URINE SEVERE")
#                             urine_list = [urine_none,urine_mild,urine_mod,urine_sev]
#                             if_none_checked_add_findings("Urine",urine_list)

#                             other_urine_none = get_radio('ctl00_ContentPlaceHolder1_RB_GEN_OSY1',"Other urine NONE")
#                             other_urine_mild = get_radio('ctl00_ContentPlaceHolder1_RB_GEN_OSY2',"other urine MILD")
#                             other_urine_mod = get_radio('ctl00_ContentPlaceHolder1_RB_GEN_OSY3',"other urine MODERATE")
#                             other_urine_sev = get_radio("ctl00_ContentPlaceHolder1_RB_GEN_OSY4","Other URINE SEVERE")
#                             other_urine_list= [other_urine_none,other_urine_mild,other_urine_mod,other_urine_sev]
#                             if_none_checked_add_findings('Other Urine', other_urine_list)

#                             urinary_continence = get_data("ctl00_ContentPlaceHolder1_Gen_URCin")
#                             urine_status = "Urinary Incontinence " + urinary_continence
#                             del visit_note_list[-1]
#                             visit_note_list.append(urine_status)

#                             insomnia_none = get_radio('ctl00_ContentPlaceHolder1_RB_SLP_INS1',"Insomnia NONE")
#                             insomnia_mild = get_radio("ctl00_ContentPlaceHolder1_RB_SLP_INS2","Insomnia MILD")
#                             insomnia_mod = get_radio("ctl00_ContentPlaceHolder1_RB_SLP_INS3","Insomnia MOD")
#                             insomnia_sev = get_radio('ctl00_ContentPlaceHolder1_RB_SLP_INS4',"insomnia SEVERE")
#                             insomnia_list = [insomnia_none,insomnia_mild,insomnia_mod,insomnia_sev]
#                             if_none_checked_add_findings("Insomnia",insomnia_list)

#                             solumn_none = get_radio('ctl00_ContentPlaceHolder1_RB_SLP_SOM1',"Solemn NONE")
#                             solumn_mild = get_radio('ctl00_ContentPlaceHolder1_RB_SLP_SOM2',"Solemn MILD")
#                             solumn_mod = get_radio('ctl00_ContentPlaceHolder1_RB_SLP_SOM3',"Solemn Moderate")
#                             solumn_sev = get_radio('ctl00_ContentPlaceHolder1_RB_SLP_SOM4',"Solemn SEVERE")
#                             solumn_list = [solumn_none,solumn_mild,solumn_mod,solumn_sev]
#                                 # if it is blank its because it wasnt checked
#                                 #printing this out to see why its not being placed on excel.
#                             if_none_checked_add_findings("Solumn",solumn_list)

#                             ############# MUSKULO ##################

#                             weakness_none = get_radio('ctl00_ContentPlaceHolder1_RB_MUS_WKN1',"Weakness NONE")
#                             weakness_mild = get_radio('ctl00_ContentPlaceHolder1_RB_MUS_WKN2',"Weakness Mild")
#                             weakness_mod = get_radio('ctl00_ContentPlaceHolder1_RB_MUS_WKN3',"Weakness Mod")
#                             weakness_sev = get_radio("ctl00_ContentPlaceHolder1_RB_MUS_WKN4","weakness Sev")
#                             weakness_list = [weakness_none,weakness_mild,weakness_mod,weakness_sev]
#                             if_none_checked_add_findings("Wekaness",weakness_list)

#                             contracture_none = get_radio('ctl00_ContentPlaceHolder1_RB_MUS_CON1',"Contracture NONE")
#                             con_mild = get_radio('ctl00_ContentPlaceHolder1_RB_MUS_CON2',"Contracture MILD")
#                             con_mod = get_radio('ctl00_ContentPlaceHolder1_RB_MUS_CON3',"Contracture MODERATE")
#                             con_severe = get_radio("ctl00_ContentPlaceHolder1_RB_MUS_CON4","Contracture Severe")
#                             contracture_list = [contracture_none,con_mild,con_mod,con_severe]
#                             if_none_checked_add_findings("Contractures",contracture_list)

#                             ########## INTEGUMENTARY #############

#                             rash_none = get_radio("ctl00_ContentPlaceHolder1_RB_INT_PRU1","rash NONE")
#                             rash_mild = get_radio('ctl00_ContentPlaceHolder1_RB_INT_PRU2',"rash MILD")
#                             rash_mod = get_radio('ctl00_ContentPlaceHolder1_RB_INT_PRU3',"rash MOD")
#                             rash_sev = get_radio('ctl00_ContentPlaceHolder1_RB_INT_PRU4',"rash SEVERE")
#                             rash_list = [rash_none,rash_mild,rash_mod,rash_sev]
#                             if_none_checked_add_findings("Rash",rash_list)

#                             wounds_none = get_radio('ctl00_ContentPlaceHolder1_RB_INT_WND1',"NO wounds")
#                             wounds_mild = get_radio('ctl00_ContentPlaceHolder1_RB_INT_WND2',"MILD wounds")
#                             wounds_mod = get_radio('ctl00_ContentPlaceHolder1_RB_INT_WND3',"MOD wounds")
#                             wounds_sev = get_radio('ctl00_ContentPlaceHolder1_RB_INT_WND4',"SEVERE wounds")
#                             wound_list = [wounds_none,wounds_mild,wounds_mod,wounds_sev]
#                             if_none_checked_add_findings("WOUNDS",wound_list)

#                             ulcers_none = get_radio('ctl00_ContentPlaceHolder1_RB_INT_OSY1',"NO ulcers")
#                             ulcers_mild = get_radio('ctl00_ContentPlaceHolder1_RB_INT_OSY2',"MILD ulcer")
#                             ulcers_mod = get_radio('ctl00_ContentPlaceHolder1_RB_INT_OSY3',"MODERATE ulcers")
#                             ulcers_sev = get_radio("ctl00_ContentPlaceHolder1_RB_INT_OSY4","SEVERE ulcers")
#                             ulcer_list = [ulcers_none,ulcers_mild,ulcers_mod,ulcers_sev]
#                             if_none_checked_add_findings("ULCERS",ulcer_list)

#                             ######## MOBILITY #################

#                             ambulatory = get_radio('ctl00_ContentPlaceHolder1_RAD_CEMO_AN',"AMBULATORY")
#                             non_ambulatory = get_radio('ctl00_ContentPlaceHolder1_RAD_CEMO_AN1',"NON-AMBULATORY")
#                             ambulatory_list = [ambulatory,non_ambulatory]
#                             if_none_checked_add_findings("Ambulation",ambulatory_list)

#                             cane = get_data('ctl00_ContentPlaceHolder1_Chk_CEMO_C')

#                             if cane == "checked": visit_note_list.append("cane")

#                             walker = get_data('ctl00_ContentPlaceHolder1_Chk_CEMO_W')
#                             if walker == "checked": visit_note_list.append("walker")

#                             wheelchair = get_data('ctl00_ContentPlaceHolder1_Chk_CEMO_WC')
#                             if wheelchair == "checked": visit_note_list.append("wheelchair")

#                             stand_by_assist = get_data('ctl00_ContentPlaceHolder1_Chk_CEMO_S')
#                             if stand_by_assist == "checked": visit_note_list.append("stand_by_assist")

#                             mod_assist = get_data('ctl00_ContentPlaceHolder1_Chk_CEMO_M')
#                             if mod_assist == "checked": visit_note_list.append("mod_assist")

#                             max_assist = get_data('ctl00_ContentPlaceHolder1_Chk_CEMO_MX')
#                             if max_assist == "checked": visit_note_list.append("MAX assist")

#                             bed_bound = get_data('ctl00_ContentPlaceHolder1_ChK_CEMO_BB')
#                             if bed_bound == "checked": visit_note_list.append("bed_bound")

#                             bed_rest = get_data('ctl00_ContentPlaceHolder1_Chk_CEMO_BR')
#                             if bed_rest == "checked": visit_note_list.append("bed_rest")

#                             up_as_tolerated = get_data('ctl00_ContentPlaceHolder1_Chk_CEMO_TO')
#                             if up_as_tolerated == "checked": visit_note_list.append("up_as_tolerated")

#                             # max_assist_non_amb = get_data('ctl00_ContentPlaceHolder1_Chk_CEMO_MXA')
#                             # if max_assist_non_amb == "checked": visit_note_list.append("up_as_tolerated")

#                             # transfer_bed = get_data('ctl00_ContentPlaceHolder1_Chk_CEMO_TC')
#                             # if transfer_bed == "checked": visit_note_list.append("transfer_bed")

#                             ############### ADL ASSESSMENT ########################


#                             ambulation = get_data('ctl00_ContentPlaceHolder1_DRP_CEAD_A')
#                             ambu_sent = "Ambulation ADL" + ambulation
#                             del visit_note_list[-1]
#                             visit_note_list.append(ambu_sent)

#                             toileting = get_data('ctl00_ContentPlaceHolder1_DRP_CEAD_C')
#                             toilet_sent = "Toilet ADL" + toileting
#                             del visit_note_list[-1]
#                             visit_note_list.append(toilet_sent)

#                             transfer_adl = get_data('ctl00_ContentPlaceHolder1_DRP_CEAD_T')
#                             transfer_sent = "Transfer ADL: "+ transfer_adl
#                             del visit_note_list[-1]  #HAD TO ADD THIS SINCE THEY ARE ADDED IN FUNCTION
#                             visit_note_list.append(transfer_sent)


#                             imminenetly_dying_yes = get_radio('ctl00_ContentPlaceHolder1_RBImminentlyDying_0',"Imminently Dying YES")
#                             imminently_dying = get_radio('ctl00_ContentPlaceHolder1_RBImminentlyDying_1',"Imminently Dying No")

#                             ################ HX OF FALLS ################

#                             falls_to_report_no = get_radio('ctl00_ContentPlaceHolder1_Env_FRA_1',"No falls to report")
#                             if falls_to_report_no == 'checked': pass
#                             else: findings.append("Check if fall report is done"+ visit_date)
#                             print("FALLS:",falls_to_report_no)

#                             kps = get_data('ctl00_ContentPlaceHolder1_DDKPS')
#                             del visit_note_list[-1]
#                             visit_note_list.append("KPS: "+kps)


#                             pps = get_data('ctl00_ContentPlaceHolder1_DDPPS')
#                             del visit_note_list[-1]
#                             visit_note_list.append("PPS: "+pps)

#                             fast = get_data("ctl00_ContentPlaceHolder1_DDFast")
#                             del visit_note_list[-1]
#                             visit_note_list.append("FAST: "+fast)

#                             nyha = get_data('ctl00_ContentPlaceHolder1_DDNYHA')
#                             del visit_note_list[-1]
#                             visit_note_list.append("NYHA: "+ nyha)

#                             ########## CARE PROVIDED #############

#                             care_provided = get_data('ctl00_ContentPlaceHolder1_CHK_CAP_PHY')
#                             structural_support = get_data('ctl00_ContentPlaceHolder1_CHK_CAP_Str')
#                             knowledge = get_data('ctl00_ContentPlaceHolder1_CHK_CAP_Kno')
#                             safety = get_data('ctl00_ContentPlaceHolder1_CHK_CAP_Saf')
#                             environmental = get_data('ctl00_ContentPlaceHolder1_CHK_CAP_Env')
#                             teaching_list = [care_provided,structural_support,knowledge,safety,environmental]

#                             for word in teaching_list:
#                                 if word != 'checked':
#                                     visit_note_list.append('Care Provided Box NOT CHECKED'+ word)
#                                     break
#                                 else:
#                                         pass

#                                             ######### VISIT CHECK LIST 333333
#                             update_family_yes = get_data('ctl00_ContentPlaceHolder1_RbVC1_0')
#                             update_cm_yes = get_data('ctl00_ContentPlaceHolder1_RbVC2_0')
#                             dme_check_yes = get_data('ctl00_ContentPlaceHolder1_RbVC4_0')
#                             check_list = [update_family_yes,update_cm_yes,dme_check_yes]


#                             for word in teaching_list:
#                                 if word != 'checked':
#                                     print("breaking")
#                                     visit_note_list.append('Care Check List not done '+ word)
#                                     break
#                             else:
#                                 pass

#                             back(2)  # THIS ONE LOOKS GOOD
#                             print("pressed back 2 good one")

#                         else:  # IT is a short form
#                             print("It is a short form")
#                             visit_note_list.append('SHORT FORM!')


#                         visit_note_list.append(visit_date)
#                         # print("VISIT NOTE LIST:",visit_note_list)
#                         # print("FINDINGS",findings)

#                     collecting()

#                     # DO NOT NEED BACK BUTTON HERE>.. this one did not go back correctly
#                     print("COLLECTING DONE..........")
#                     append_to_excel(visit_note_list)


#             print("GOING back ")
#             back(2)
#             print("didd it go back correctly to the first row page?")

#             sound()
#             sleep(30)
#             print("Pressed back 2 Z.. is it correct?")


#     else:
#         def one_page_visit_check():
#             print("*******************one_page_visit_check*********************")
#             page_num = 1
#             click(f'/html/body/form/table/tbody/tr[1]/td/div/table/tbody/tr[2]/td/table/tbody/tr[4]/td[2]/div/table/tbody/tr/td[2]/div[2]/div/table/tbody/tr[17]/td/table/tbody/tr/td[{page_num}]' )
#             print("clicked pageb",page_num)
#             page_num +=1

#             link_counter = 2 #starting 2 since xpath starts w 2
#             date_note_list = []
#             x_num_list = []   #gotta clean list each iteration
#             x_num =2   #

#             visit_note_rows = find_elements('/html/body/form/table/tbody/tr[1]/td/div/table/tbody/tr[2]/td/table/tbody/tr[4]/td[2]/div/table/tbody/tr/td[2]/div[2]/div/table/tbody/tr[*]')
#             print("LEN VISIT NOTE ROWS B: ",len(visit_note_rows))
#                         ###### SLICED BECAUSE CAN'T USE FIRST AND LAST
#             for row in visit_note_rows[1:-1]:
#                 row_text = row.text
#                 print(row_text)

#                 if "RN" in row_text or "LVN" in row_text or "BSN" in row_text:

#                     x_num_list.append(x_num)
#                     print("X_NUM:",x_num)
#                 else:
#                     print("no num")
#                 x_num+= 1
#                 # print("x_num_list",x_num_list)
#                     ##NOW CHANGE XPATH FROM THE LIST ITERATION. AND GATHER FROM PAGES AND ALL

#             #CLICKS ON VISIT NOTE TO GATHER DATA
#                 #added this to take out back button at bottom of loop

#             for num in x_num_list:  #
#                 print("Starting to click visit_nurse rowb:",num)
#                 visit_note_list = []  #should collect everything then refresh it
#                 visit_nurse_row = f'/html/body/form/table/tbody/tr[1]/td/div/table/tbody/tr[2]/td/table/tbody/tr[4]/td[2]/div/table/tbody/tr/td[2]/div[2]/div/table/tbody/tr[{num}]/td[1]/input'
#                 click(visit_nurse_row)
#                 # print("NUM2:",x_num)
#                 print('Reload page to get sourceb')
#                 # expand = click('//*[@id="ctl00_ContentPlaceHolder1_ImgExpandAll"]')
#                 # sleep(8)
#                 url1 = driver.current_url
#                 new = url1[:-3]+ 'c=&drl=exp'

#                 driver.get(new)

#                 visit_note2 = driver.page_source

#                 soup = BeautifulSoup(visit_note2, 'html.parser')

#                 comp_filename2 = "visit note"

#                 with open(comp_filename2, 'wt', encoding='utf-8') as html_file:  #writes code name
#                     for line in soup.prettify():
#                             html_file.write(line)

#                 print("file Visit Note Sourcepage saved as: ",comp_filename2)  # ADDED THE 2 on this

#                 with open(comp_filename2, 'r',encoding='utf-8') as file:  # DO I NEED THIS
#                     soup = BeautifulSoup(file, 'html.parser')

#                     def get_data(id):
#                         try:
#                             item = soup.find(id=re.compile(id))
#                             if 'type="checkbox"' in str(item):
#                                     if 'checked="checked"' in str(item):
#                                         return 'checked'
#                                     else:
#                                         return "not checked"

#                             selected = item.select_one('option:checked')
#                             value = selected.get_text("|", strip=True)
#                         except:
#                             try:
#                                     item = soup.find(id=re.compile(id))
#                                     value = item['value']

#                             except:
#                                     item = soup.find(id=re.compile(id))
#                                     value = item.get_text("|", strip=True)

#                         if value == '1' or value == '2':
#                             if 'checked="checked"' in str(item):
#                                     # print('checked')
#                                     return 'checked'

#                         visit_note_list.append(value)
#                         # print(value)
#                         return value


#                     def get_radio(id,print_name):
#                         item = soup.find(id=re.compile(id))
#                         if 'checked="checked"' in str(item):
#                             name = print_name
#                             visit_note_list.append(name)
#                             print(name)
#                             value = 'checked'
#                         else:
#                             value = ''

#                         return value

#                     def get_text_value(id):
#                         item = soup.find(id=re.compile(id))
#                         value = item['value']
#                         print("Value")
#                         return value

#                     def collecting():  #added this just so i can keep it all together

#                         visit_date = get_text_value('ctl00_ContentPlaceHolder1_txtHDate')
#                         visit_note_list.append(visit_date)

#                         # del visit_note_list[0]
#                         nurse = get_data('ctl00_ContentPlaceHolder1_DropCreatedBy')

#                         # nurse = get_data('ctl00_ContentPlaceHolder1_DropCreatedBy')
#                         discipline = get_data('ctl00_ContentPlaceHolder1_DropDiscipline')
#                         care_level = get_data('ctl00_ContentPlaceHolder1_DropCareLevel')

#                         narrative = soup.find(id=re.compile('ctl00_ContentPlaceHolder1_TxtComments'))
#                         narrative = narrative.get_text("|", strip=True)
#                         narrative = add_char_every_certain_num_char_in_string(50,narrative)

#                         if len(narrative) < 5:
#                             findings.append("Missing Nurse Narrative by "+ nurse + ' '+visit_date)
#                         visit_note_list.append(narrative)
#                         # print("Narrative",narrative)

#                         # narrative = narrative.get_text("|", strip=True)
#                         # print(narrative)


#                         # del visit_note_list[-1]


#                         # frequency = get_data('ctl00_ContentPlaceHolder1_lblFrequencyOfVisit')
#                         chief_complaint = get_data('ctl00_ContentPlaceHolder1_TxtPsgStatesChiefComplaint')

#                         ############## COVID #########################
#                         covid_screen_yes = get_data('ctl00_ContentPlaceHolder1_rbPriorScreening_1')

#                         # del visit_note_list[-1]
#                         if covid_screen_yes == 'checked':
#                             pass

#                         else:  findings.append('Covid Screen not done. ')

#                         test_date = get_data('ctl00_ContentPlaceHolder1_txtPriorScreenDate')
#                         # del visit_note_list[-1]
#                         if test_date == '': findings.append("Covid test date not entered" + visit_date)

#                         pt_report_positive = get_data('ctl00_ContentPlaceHolder1_rbPositive_0')
#                         # del visit_note_list[-1]
#                         if pt_report_positive != 'checked': findings.append("pt reported positive"+visit_date)

#                         pt_report_fever = get_data('ctl00_ContentPlaceHolder1_rbCoronaVirus_0')
#                         # del visit_note_list[-1]
#                         if pt_report_positive != 'checked': findings.append("pt report fever"+visit_date)

#                         add_comment = get_data('ctl00_ContentPlaceHolder1_rbPatientPcgVisit_0')
#                         # del visit_note_list[-1]
#                         if pt_report_positive != 'checked': findings.append("additional comment in covid")


#                         followed_covid = get_data('ctl00_ContentPlaceHolder1_chkCoronaInstructions')
#                         del visit_note_list[-1]
#                         if followed_covid != 'checked': findings.append("pt reported covid positive")

#                         utilized_ppe = get_data('ctl00_ContentPlaceHolder1_chkPpe')
#                         del visit_note_list[-1]
#                         if utilized_ppe != 'checked': findings.append("nurse didn't select box utilized ppe")

#                         ################################# PPE######################################

#                         is_pain_controlled = get_data('ctl00_ContentPlaceHolder1_RbControlled_0')
#                         # del visit_note_list[-1]
#                         if is_pain_controlled != 'Yes': findings.append('Check pain assessment since pt pain not controlled')

#                         pain_reported = get_data('ctl00_ContentPlaceHolder1_DropPainLevel')
#                         # del visit_note_list[-1]
#                         visit_note_list.append("Pain: "+pain_reported)

#                         ##################### SIGNS AND SYMPTOMS #############################

#                         ################ NEURO ###############
#                         #PRINTS OUT THE ACTUAL SELECTED ONE

#                         def if_none_checked_add_findings(category,list):
#                             num = 0
#                             for var in list:
#                                 if var == 'checked':
#                                         num += 1
#                                         pass
#                                 else:
#                                         pass
#                             if num == 0:
#                                 findings.append(category + ' boxes were not selected in visit notes')


#                         form_type = get_data('ctl00_ContentPlaceHolder1_DrpVisitUnsc')
#                         del visit_note_list[-1]
#                         visit_note_list.append('Form Type2:'+form_type)

#                         #STILL NEED TO ADD TEXT INPUTS TO THESE
#                         if form_type != 'Short Form':
#                             print("Form_type:",form_type)

#                             anxiety = get_radio('ctl00_ContentPlaceHolder1_RB_NMS_Anx1','anxiety none')
#                             anxiety_mild = get_radio('ctl00_ContentPlaceHolder1_RB_NMS_Anx2','anxiety_mild')
#                             anxiety_mod = get_radio('ctl00_ContentPlaceHolder1_RB_NMS_Anx3','anxiety_mod')
#                             anxiety_severe = get_radio('ctl00_ContentPlaceHolder1_RB_NMS_Anx4','anxiety_severe')
#                             anxiety_list = [anxiety,anxiety_mild,anxiety_mod,anxiety_severe]
#                             if_none_checked_add_findings("Anxiety",anxiety_list)

#                             depression_none= get_radio('ctl00_ContentPlaceHolder1_RB_NMS_Dep1','depression_none')
#                             dep_mild = get_radio('ctl00_ContentPlaceHolder1_RB_NMS_Dep2','dep_mild')
#                             dep_mod = get_radio('ctl00_ContentPlaceHolder1_RB_NMS_Dep3','dep_mod')
#                             dep_severe = get_radio('ctl00_ContentPlaceHolder1_RB_NMS_Dep4','dep_severe')
#                             depression_list = [depression_none,dep_mild,dep_mod,dep_severe]
#                             if_none_checked_add_findings("Depression",depression_list)

#                             agitation_none = get_radio('ctl00_ContentPlaceHolder1_RB_NMS_DIZ1','agitation_none')
#                             agitation_mild = get_radio('ctl00_ContentPlaceHolder1_RB_NMS_DIZ2','agitation_mild')
#                             agitation_mod = get_radio('ctl00_ContentPlaceHolder1_RB_NMS_DIZ3','agitation_mod')
#                             agitation_severe = get_radio('ctl00_ContentPlaceHolder1_RB_NMS_DIZ4','agitation_severe')
#                             agitation_list = [agitation_none,agitation_mild,agitation_mod,agitation_severe]
#                             if_none_checked_add_findings("Agitation",agitation_list)


#                             confusion_none = get_radio('ctl00_ContentPlaceHolder1_RBS_NMS_CON1','confusion_none')
#                             confusion_mild = get_radio('ctl00_ContentPlaceHolder1_RBS_NMS_CON2','confusion_mild')
#                             confusion_mod = get_radio('ctl00_ContentPlaceHolder1_RBS_NMS_CON3','confusion_mod')
#                             confusion_sev = get_radio('ctl00_ContentPlaceHolder1_RBS_NMS_CON4','confusion_sev')
#                             confusion_list = [confusion_none,confusion_mild,confusion_mod,confusion_sev]
#                             if_none_checked_add_findings("confusion",confusion_list)

#                             speech_impair_none = get_radio('ctl00_ContentPlaceHolder1_RBS_SI_CON1','speech_impair_none')
#                             speech_impair_mild = get_radio('ctl00_ContentPlaceHolder1_RBS_SI_CON2','speech_impair_mild')
#                             speech_impair_mod = get_radio('ctl00_ContentPlaceHolder1_RBS_SI_CON3','speech_impair_mod')
#                             speech_impair_sev = get_radio('ctl00_ContentPlaceHolder1_RBS_SI_CON4','speech_impair_sev')
#                             speech_list = [speech_impair_none,speech_impair_mild,speech_impair_mod,speech_impair_sev]
#                             if_none_checked_add_findings("Speech",speech_list)

#                             other_symp_none = get_radio('ctl00_ContentPlaceHolder1_RB_NMS_OSY1','other_sym_none')
#                             other_mild = get_radio('ctl00_ContentPlaceHolder1_RB_NMS_OSY2','other_sym_mild')
#                             other_mod = get_radio('ctl00_ContentPlaceHolder1_RB_NMS_OSY3','other_symp_mod')
#                             other_sev = get_radio('ctl00_ContentPlaceHolder1_RB_NMS_OSY4','other_symp_sev')
#                             other_list = [other_symp_none,other_mild,other_mod,other_sev]
#                             if_none_checked_add_findings('Other Symptoms',other_list)

#                             #################### CARDIO#################################

#                             arrhythmia_none = get_radio('ctl00_ContentPlaceHolder1_RB_CAR_ARR1', "arrhythmia_none")
#                             arr_mild = get_radio("ctl00_ContentPlaceHolder1_RB_CAR_ARR2","Arrythmia Mild")
#                             arr_mod = get_radio('ctl00_ContentPlaceHolder1_RB_CAR_ARR3','Arrythmia mod')
#                             arr_sev = get_radio('ctl00_ContentPlaceHolder1_RB_CAR_ARR4',"Arrhythmia Severe")
#                             arrhythmia_list = [arrhythmia_none,arr_mild,arr_mod,arr_sev]
#                             if_none_checked_add_findings("Arrhythmia", arrhythmia_list)

#                             edema_none = get_radio('ctl00_ContentPlaceHolder1_RB_CAR_EDM1','Edema None')
#                             edema_mild = get_radio('ctl00_ContentPlaceHolder1_RB_CAR_EDM2',"Edema Mild")
#                             edema_mod = get_radio('ctl00_ContentPlaceHolder1_RB_CAR_EDM3','Edema Mod')
#                             edema_sev = get_radio('ctl00_ContentPlaceHolder1_RB_CAR_EDM4',"Edema Severe")
#                             edema_list = [edema_none,edema_mild,edema_mod,edema_sev]
#                             if_none_checked_add_findings("EDEMA", edema_list)

#                             chest_pain_none = get_radio('ctl00_ContentPlaceHolder1_RB_CAR_CHP1',"No Chest Pain")
#                             chest_pain_mild = get_radio('ctl00_ContentPlaceHolder1_RB_CAR_CHP2',"Mild chest pain")
#                             chest_pain_mod = get_radio('ctl00_ContentPlaceHolder1_RB_CAR_CHP3',"Moderate chest pain")
#                             chest_pain_sev = get_radio('ctl00_ContentPlaceHolder1_RB_CAR_CHP4',"severe chest pain")
#                             chest_pain_oth = get_radio('ctl00_ContentPlaceHolder1_RB_CAR_OSY1',"chest pain other")
#                             chest_pain_list = [chest_pain_none,chest_pain_mild,chest_pain_mod,chest_pain_sev,chest_pain_oth]
#                             if_none_checked_add_findings("Chest Pain",chest_pain_list)

#                             ################# RESPIRATORY #######################
#                             dyspnea_none = get_radio('ctl00_ContentPlaceHolder1_RB_CAR_DYS1',"Dyspnea None")
#                             dyspnea_mild = get_radio('ctl00_ContentPlaceHolder1_RB_CAR_DYS2',"Dyspnea Mild")
#                             dyspnea_mod = get_radio('ctl00_ContentPlaceHolder1_RB_CAR_DYS3',"Dyspnea Mod")
#                             dyspnea_sev = get_radio('ctl00_ContentPlaceHolder1_RB_CAR_DYS4',"Dyspnea Sev")
#                             dyspnea_list = [dyspnea_none,dyspnea_mild,dyspnea_mod,dyspnea_sev]
#                             if_none_checked_add_findings("Dyspnea",dyspnea_list)

#                             cough_none = get_radio('ctl00_ContentPlaceHolder1_RB_C1',"Cough None")
#                             cough_mild = get_radio('ctl00_ContentPlaceHolder1_RB_C2',"Cough Mild")
#                             cough_mod = get_radio('ctl00_ContentPlaceHolder1_RB_C3',"Cough Mod")
#                             cough_sev = get_radio('ctl00_ContentPlaceHolder1_RB_C4',"Cough Severe")
#                             cough_list = [cough_none,cough_mild,cough_mod,cough_sev]
#                             if_none_checked_add_findings('Cough',cough_list)

#                             abnormal_lung_sounds = get_radio('ctl00_ContentPlaceHolder1_RB_A1', "Abnormal Lung Sounds None")
#                             lung_sounds_mild = get_radio('ctl00_ContentPlaceHolder1_RB_A2',"Abnormal Lung Sounds Mild")
#                             lung_sounds_mod = get_radio('ctl00_ContentPlaceHolder1_RB_A3',"Abnormal Lung Sounds Mild")
#                             lung_sounds_sev = get_radio('ctl00_ContentPlaceHolder1_RB_A4',"Abnormal lung sounds Severe")
#                             lung_sound_list = [abnormal_lung_sounds,lung_sounds_mild,lung_sounds_mod,lung_sounds_sev]
#                             if_none_checked_add_findings('Lung Sounds',lung_sound_list)

#                             lung_sound_other_none = get_radio('ctl00_ContentPlaceHolder1_RB_RES_OSY1',"Lung Sounds Other NONE")
#                             ls_other_mild = get_radio('ctl00_ContentPlaceHolder1_RB_RES_OSY2', "Lung Sounds other MILD")
#                             ls_other_mod = get_radio('ctl00_ContentPlaceHolder1_RB_RES_OSY3',"Lung Sounds other MOD")
#                             ls_other_sev = get_radio('ctl00_ContentPlaceHolder1_RB_RES_OSY4',"Lung Sounds other SEVERE")
#                             ls_other_list = [lung_sound_other_none,ls_other_mild,ls_other_mod,ls_other_sev]
#                             if_none_checked_add_findings("Lung Sounds other",ls_other_list)

#                             ################ IMMUNOLOGICAL ################################
#                             infection_none = get_radio('ctl00_ContentPlaceHolder1_RB_IMG_INF1',"INFECTION NONE")
#                             infection_mild = get_radio('ctl00_ContentPlaceHolder1_RB_IMG_INF2',"Infection MILD")
#                             infection_mod = get_radio('ctl00_ContentPlaceHolder1_RB_IMG_INF3',"Infection MOD")
#                             infection_sev = get_radio('ctl00_ContentPlaceHolder1_RB_IMG_INF4',"Infection SEVERE")
#                             infection_list = [infection_none,infection_mild,infection_mod,infection_sev]
#                             if_none_checked_add_findings('Infections',infection_list)

#                             ############### GASTRO INTESTINAL ###############################

#                             nausea_none = get_radio('ctl00_ContentPlaceHolder1_RB_GAS_NAU1',"Nausea None")
#                             nausea_mild = get_radio('ctl00_ContentPlaceHolder1_RB_GAS_NAU2',"Nausea MILD")
#                             nausea_mod = get_radio('ctl00_ContentPlaceHolder1_RB_GAS_NAU3',"Nausea MOD")
#                             nausea_sev = get_radio('ctl00_ContentPlaceHolder1_RB_GAS_NAU4',"Nausea SEV")
#                             nausea_list = [nausea_none,nausea_mild,nausea_mod,nausea_sev]
#                             if_none_checked_add_findings("Nausea",nausea_list)

#                             vomiting_none = get_radio('ctl00_ContentPlaceHolder1_RB_GAS_VOM1',"vomiting_none")
#                             vomit_mild = get_radio('ctl00_ContentPlaceHolder1_RB_GAS_VOM2',"vomit_mild")
#                             vomit_mod = get_radio('ctl00_ContentPlaceHolder1_RB_GAS_VOM3',"vomit_mod")
#                             vomit_sev = get_radio('ctl00_ContentPlaceHolder1_RB_GAS_VOM4',"vomit_sev")
#                             vomit_list = [vomiting_none,vomit_mild,vomit_mod,vomit_sev]
#                             if_none_checked_add_findings("Vomitting", vomit_list)

#                             constipation_none = get_radio('ctl00_ContentPlaceHolder1_RB_GAS_CON1', "Constipation NONE")
#                             con_mild = get_radio('ctl00_ContentPlaceHolder1_RB_GAS_CON2','Constipation MILD')
#                             con_mod = get_radio('ctl00_ContentPlaceHolder1_RB_GAS_CON3',"Constipation MOD")
#                             con_sev = get_radio('ctl00_ContentPlaceHolder1_RB_GAS_CON4',"Constipation SEVERE")
#                             constipation_list = [constipation_none,con_mild,con_mod,con_sev]
#                             if_none_checked_add_findings("Constipation",constipation_list)

#                             diarrhea_none = get_radio('ctl00_ContentPlaceHolder1_RB_GAS_DIA1',"Diarrhea NONE")
#                             diarrhea_mild = get_radio('ctl00_ContentPlaceHolder1_RB_GAS_DIA2',"Diarrhea MILD")
#                             diarrhea_mod = get_radio('ctl00_ContentPlaceHolder1_RB_GAS_DIA3',"Diarrhea MODERATE")
#                             diarrhea_sev = get_radio('ctl00_ContentPlaceHolder1_RB_GAS_DIA4',"Diarrhea SEVERE")
#                             diarrhea_list = [diarrhea_none,diarrhea_mild,diarrhea_mod,diarrhea_sev]
#                             if_none_checked_add_findings("Diarrhea",diarrhea_list)

#                             other_gi_none = get_radio('ctl00_ContentPlaceHolder1_RB_GAS_DIA1',"Other GI NONE")
#                             other_gi_mild = get_radio('ctl00_ContentPlaceHolder1_RB_GAS_DIA2',"Other GI Mild")
#                             other_gi_mod = get_radio('ctl00_ContentPlaceHolder1_RB_GAS_DIA3',"other GI Mod")
#                             other_gi_sev = get_radio('ctl00_ContentPlaceHolder1_RB_GAS_DIA4',"Other GI SEVERE")
#                             other_gi_list = [other_gi_none,other_gi_mild,other_gi_mod,other_gi_sev]
#                             if_none_checked_add_findings("Other GI",other_gi_list)

#                             last_bm = get_data("ctl00_ContentPlaceHolder1_LastBM")
#                             bm_last = "Last BM: " + last_bm
#                             del visit_note_list[-1]
#                             visit_note_list.append(bm_last)

#                             incontinent = get_data('ctl00_ContentPlaceHolder1_Gas_STS')
#                             incontinent_sent = 'Incontinent: '+ incontinent
#                             del visit_note_list[-1]
#                             visit_note_list.append(incontinent_sent)

#                             ############## NUTRITION #################

#                             perc_intake_none = get_radio("ctl00_ContentPlaceHolder1_RB_GAS_OI1", "Perc Oral Intake NONE")
#                             perc_intake_mild = get_radio('ctl00_ContentPlaceHolder1_RB_GAS_OI2',"Perc Oral Intake Mild")
#                             perc_intake_mod = get_radio("ctl00_ContentPlaceHolder1_RB_GAS_OI3","perc Oral Intake Mod")
#                             perc_intake_sev = get_radio('ctl00_ContentPlaceHolder1_RB_GAS_OI4',"Perc intake SEVERE")
#                             intake_list = [perc_intake_none,perc_intake_mild,perc_intake_mod,perc_intake_sev]
#                             if_none_checked_add_findings("Percentage of Oral Intake",intake_list)

#                             nutrition_other_none = get_radio("ctl00_ContentPlaceHolder1_RB_NUT_OSY1","Nutrition Other NONE")
#                             nut_other_mild = get_radio('ctl00_ContentPlaceHolder1_RB_NUT_OSY2',"Nutrition Other MILD")
#                             nut_other_mod = get_radio('ctl00_ContentPlaceHolder1_RB_NUT_OSY3',"Nutrition Other MOD")
#                             nut_other_sev = get_radio('ctl00_ContentPlaceHolder1_RB_NUT_OSY4',"Nutrition other SEVERE")
#                             nut_other_list = [nutrition_other_none,nut_other_mild,nut_other_mod,nut_other_sev]
#                             if_none_checked_add_findings("Nutrition Other",nut_other_list)

#                             artificially_fed_no = get_radio("ctl00_ContentPlaceHolder1_Gas_AFF_0","Artifically fed No")
#                             fed_peg = get_radio('ctl00_ContentPlaceHolder1_Gas_AFF_1',"PEG Feeding")
#                             fed_ng = get_radio('ctl00_ContentPlaceHolder1_Gas_AFF_2',"NG Tube")
#                             fed_j_tube = get_radio('ctl00_ContentPlaceHolder1_Gas_AFF_3',"J-Tube")
#                             fed_pump = get_radio("ctl00_ContentPlaceHolder1_Gas_AFF_4","Pump Feeding")
#                             fed_tpn = get_radio('ctl00_ContentPlaceHolder1_Gas_AFF_5',"TPN Feeding")
#                             art_fed_list = [artificially_fed_no,fed_peg,fed_j_tube,fed_pump,fed_tpn]
#                             if_none_checked_add_findings("Artificially fed",art_fed_list)
#                             diet = get_data("ctl00_ContentPlaceHolder1_DrpDietType")

#                             ########### ENDOCRINE #####################
#                             bs_none = get_radio('ctl00_ContentPlaceHolder1_RB_IND_DIA1',"Blood Sugar NONE")
#                             bs_mild = get_radio('ctl00_ContentPlaceHolder1_RB_IND_DIA2',"Mild Blood Sugar")
#                             bs_mod = get_radio('ctl00_ContentPlaceHolder1_RB_IND_DIA3',"Mod Blood Sugar")
#                             bs_sev = get_radio("ctl00_ContentPlaceHolder1_RB_IND_DIA4","Severe BS")
#                             bs_list = [bs_none,bs_mild,bs_mod,bs_sev]
#                             if_none_checked_add_findings("Blood Sugar",bs_list)

#                             other_bs_none = get_radio('ctl00_ContentPlaceHolder1_RB_IND_OSY1',"other BS NONE")
#                             other_bs_mild = get_radio('ctl00_ContentPlaceHolder1_RB_IND_OSY2',"other BS MILD")
#                             other_bs_mod = get_radio('ctl00_ContentPlaceHolder1_RB_IND_OSY3',"Other BS MOD")
#                             other_bs_sev = get_radio('ctl00_ContentPlaceHolder1_RB_IND_OSY4',"Other BS SEVERE")
#                             other_bs_list = [other_bs_none,other_bs_mild,other_bs_mod,other_bs_sev]
#                             if_none_checked_add_findings("Other BS",other_bs_list)

#                             ####### GENITOURINARY REPRODUCTIVE ###########

#                             urine_none = get_radio('ctl00_ContentPlaceHolder1_RB_GEN_URP1',"Urine No issues")
#                             urine_mild = get_radio('ctl00_ContentPlaceHolder1_RB_GEN_URP2',"Urine MILD")
#                             urine_mod = get_radio('ctl00_ContentPlaceHolder1_RB_GEN_URP3',"Urine MODERATE")
#                             urine_sev = get_radio('ctl00_ContentPlaceHolder1_RB_GEN_URP4',"URINE SEVERE")
#                             urine_list = [urine_none,urine_mild,urine_mod,urine_sev]
#                             if_none_checked_add_findings("Urine",urine_list)

#                             other_urine_none = get_radio('ctl00_ContentPlaceHolder1_RB_GEN_OSY1',"Other urine NONE")
#                             other_urine_mild = get_radio('ctl00_ContentPlaceHolder1_RB_GEN_OSY2',"other urine MILD")
#                             other_urine_mod = get_radio('ctl00_ContentPlaceHolder1_RB_GEN_OSY3',"other urine MODERATE")
#                             other_urine_sev = get_radio("ctl00_ContentPlaceHolder1_RB_GEN_OSY4","Other URINE SEVERE")
#                             other_urine_list= [other_urine_none,other_urine_mild,other_urine_mod,other_urine_sev]
#                             if_none_checked_add_findings('Other Urine', other_urine_list)

#                             urinary_continence = get_data("ctl00_ContentPlaceHolder1_Gen_URCin")
#                             urine_status = "Urinary Incontinence " + urinary_continence
#                             del visit_note_list[-1]
#                             visit_note_list.append(urine_status)

#                             insomnia_none = get_radio('ctl00_ContentPlaceHolder1_RB_SLP_INS1',"Insomnia NONE")
#                             insomnia_mild = get_radio("ctl00_ContentPlaceHolder1_RB_SLP_INS2","Insomnia MILD")
#                             insomnia_mod = get_radio("ctl00_ContentPlaceHolder1_RB_SLP_INS3","Insomnia MOD")
#                             insomnia_sev = get_radio('ctl00_ContentPlaceHolder1_RB_SLP_INS4',"insomnia SEVERE")
#                             insomnia_list = [insomnia_none,insomnia_mild,insomnia_mod,insomnia_sev]
#                             if_none_checked_add_findings("Insomnia",insomnia_list)

#                             ############# MUSKULO ##################

#                             weakness_none = get_radio('ctl00_ContentPlaceHolder1_RB_MUS_WKN1',"Weakness NONE")
#                             weakness_mild = get_radio('ctl00_ContentPlaceHolder1_RB_MUS_WKN2',"Weakness Mild")
#                             weakness_mod = get_radio('ctl00_ContentPlaceHolder1_RB_MUS_WKN3',"Weakness Mod")
#                             weakness_sev = get_radio("ctl00_ContentPlaceHolder1_RB_MUS_WKN4","weakness Sev")
#                             weakness_list = [weakness_none,weakness_mild,weakness_mod,weakness_sev]
#                             if_none_checked_add_findings("Wekaness",weakness_list)

#                             contracture_none = get_radio('ctl00_ContentPlaceHolder1_RB_MUS_CON1',"Contracture NONE")
#                             con_mild = get_radio('ctl00_ContentPlaceHolder1_RB_MUS_CON2',"Contracture MILD")
#                             con_mod = get_radio('ctl00_ContentPlaceHolder1_RB_MUS_CON3',"Contracture MODERATE")
#                             con_severe = get_radio("ctl00_ContentPlaceHolder1_RB_MUS_CON4","Contracture Severe")
#                             contracture_list = [contracture_none,con_mild,con_mod,con_severe]
#                             if_none_checked_add_findings("Contractures",contracture_list)

#                             ########## INTEGUMENTARY #############

#                             rash_none = get_radio("ctl00_ContentPlaceHolder1_RB_INT_PRU1","rash NONE")
#                             rash_mild = get_radio('ctl00_ContentPlaceHolder1_RB_INT_PRU2',"rash MILD")
#                             rash_mod = get_radio('ctl00_ContentPlaceHolder1_RB_INT_PRU3',"rash MOD")
#                             rash_sev = get_radio('ctl00_ContentPlaceHolder1_RB_INT_PRU4',"rash SEVERE")
#                             rash_list = [rash_none,rash_mild,rash_mod,rash_sev]
#                             if_none_checked_add_findings("Rash",rash_list)

#                             wounds_none = get_radio('ctl00_ContentPlaceHolder1_RB_INT_WND1',"NO wounds")
#                             wounds_mild = get_radio('ctl00_ContentPlaceHolder1_RB_INT_WND2',"MILD wounds")
#                             wounds_mod = get_radio('ctl00_ContentPlaceHolder1_RB_INT_WND3',"MOD wounds")
#                             wounds_sev = get_radio('ctl00_ContentPlaceHolder1_RB_INT_WND4',"SEVERE wounds")
#                             wound_list = [wounds_none,wounds_mild,wounds_mod,wounds_sev]
#                             if_none_checked_add_findings("WOUNDS",wound_list)

#                             ulcers_none = get_radio('ctl00_ContentPlaceHolder1_RB_INT_OSY1',"NO ulcers")
#                             ulcers_mild = get_radio('ctl00_ContentPlaceHolder1_RB_INT_OSY2',"MILD ulcer")
#                             ulcers_mod = get_radio('ctl00_ContentPlaceHolder1_RB_INT_OSY3',"MODERATE ulcers")
#                             ulcers_sev = get_radio("ctl00_ContentPlaceHolder1_RB_INT_OSY4","SEVERE ulcers")
#                             ulcer_list = [ulcers_none,ulcers_mild,ulcers_mod,ulcers_sev]
#                             if_none_checked_add_findings("ULCERS",ulcer_list)

#                             ######## MOBILITY #################

#                             ambulatory = get_radio('ctl00_ContentPlaceHolder1_RAD_CEMO_AN',"AMBULATORY")
#                             non_ambulatory = get_radio('ctl00_ContentPlaceHolder1_RAD_CEMO_AN1',"NON-AMBULATORY")
#                             ambulatory_list = [ambulatory,non_ambulatory]
#                             if_none_checked_add_findings("Ambulation",ambulatory_list)

#                             cane = get_data('ctl00_ContentPlaceHolder1_Chk_CEMO_C')

#                             if cane == "checked": visit_note_list.append("cane")

#                             walker = get_data('ctl00_ContentPlaceHolder1_Chk_CEMO_W')
#                             if walker == "checked": visit_note_list.append("walker")

#                             wheelchair = get_data('ctl00_ContentPlaceHolder1_Chk_CEMO_WC')
#                             if wheelchair == "checked": visit_note_list.append("wheelchair")

#                             stand_by_assist = get_data('ctl00_ContentPlaceHolder1_Chk_CEMO_S')
#                             if stand_by_assist == "checked": visit_note_list.append("stand_by_assist")

#                             mod_assist = get_data('ctl00_ContentPlaceHolder1_Chk_CEMO_M')
#                             if mod_assist == "checked": visit_note_list.append("mod_assist")

#                             max_assist = get_data('ctl00_ContentPlaceHolder1_Chk_CEMO_MX')
#                             if max_assist == "checked": visit_note_list.append("MAX assist")

#                             bed_bound = get_data('ctl00_ContentPlaceHolder1_ChK_CEMO_BB')
#                             if bed_bound == "checked": visit_note_list.append("bed_bound")

#                             bed_rest = get_data('ctl00_ContentPlaceHolder1_Chk_CEMO_BR')
#                             if bed_rest == "checked": visit_note_list.append("bed_rest")

#                             up_as_tolerated = get_data('ctl00_ContentPlaceHolder1_Chk_CEMO_TO')
#                             if up_as_tolerated == "checked": visit_note_list.append("up_as_tolerated")

#                             # max_assist_non_amb = get_data('ctl00_ContentPlaceHolder1_Chk_CEMO_MXA')
#                             # if max_assist_non_amb == "checked": visit_note_list.append("up_as_tolerated")

#                             # transfer_bed = get_data('ctl00_ContentPlaceHolder1_Chk_CEMO_TC')
#                             # if transfer_bed == "checked": visit_note_list.append("transfer_bed")

#                             ############### ADL ASSESSMENT ########################

#                             ambulation = get_data('ctl00_ContentPlaceHolder1_DRP_CEAD_A')
#                             ambu_sent = "Ambulation ADL" + ambulation
#                             del visit_note_list[-1]
#                             visit_note_list.append(ambu_sent)

#                             toileting = get_data('ctl00_ContentPlaceHolder1_DRP_CEAD_C')
#                             toilet_sent = "Toilet ADL: " + toileting
#                             del visit_note_list[-1]
#                             visit_note_list.append(toilet_sent)

#                             transfer_adl = get_data('ctl00_ContentPlaceHolder1_DRP_CEAD_T')
#                             transfer_sent = "Transfer ADL: "+ transfer_adl
#                             del visit_note_list[-1]  #HAD TO ADD THIS SINCE THEY ARE ADDED IN FUNCTION
#                             visit_note_list.append(transfer_sent)

#                             total_adl_score = 'ctl00_ContentPlaceHolder1_TotalADL'
#                             item = soup.find(id=re.compile(total_adl_score))
#                             total_adl = item.get_text("|", strip=True)
#                             visit_note_list.append("Total ADL Score: "+total_adl)

#                             print("TOTAL_ADL_SCORE",total_adl)

#                             imminenetly_dying_yes = get_radio('ctl00_ContentPlaceHolder1_RBImminentlyDying_0',"Imminently Dying YES")
#                             imminently_dying = get_radio('ctl00_ContentPlaceHolder1_RBImminentlyDying_1',"Imminently Dying No")

#                             ################ HX OF FALLS ################

#                             falls_to_report_no = get_radio('ctl00_ContentPlaceHolder1_Env_FRA_1',"No falls to report")
#                             if falls_to_report_no == 'checked': pass
#                             else: findings.append("They need to report Fall, make sure report was done")
#                             print("FALLS:",falls_to_report_no)

#                             kps = get_data('ctl00_ContentPlaceHolder1_DDKPS')
#                             del visit_note_list[-1]
#                             visit_note_list.append("KPS: "+kps)

#                             pps = get_data('ctl00_ContentPlaceHolder1_DDPPS')
#                             del visit_note_list[-1]
#                             visit_note_list.append("PPS: "+pps)

#                             fast = get_data("ctl00_ContentPlaceHolder1_DDFast")
#                             del visit_note_list[-1]
#                             visit_note_list.append("FAST: "+fast)

#                             nyha = get_data('ctl00_ContentPlaceHolder1_DDNYHA')
#                             del visit_note_list[-1]
#                             visit_note_list.append("NYHA: "+ nyha)

#                             ########## CARE PROVIDED #############

#                             care_provided = get_data('ctl00_ContentPlaceHolder1_CHK_CAP_PHY')
#                             structural_support = get_data('ctl00_ContentPlaceHolder1_CHK_CAP_Str')
#                             knowledge = get_data('ctl00_ContentPlaceHolder1_CHK_CAP_Kno')
#                             safety = get_data('ctl00_ContentPlaceHolder1_CHK_CAP_Saf')
#                             environmental = get_data('ctl00_ContentPlaceHolder1_CHK_CAP_Env')
#                             teaching_list = [care_provided,structural_support,knowledge,safety,environmental]
#                             for word in teaching_list:
#                                 if word != 'checked':
#                                     visit_note_list.append('Care Provided Box NOT CHECKED'+ word)
#                                     break
#                                 else:
#                                     pass

#                                             ######### VISIT CHECK LIST 333333

#                             update_family_yes = get_data('ctl00_ContentPlaceHolder1_RbVC1_0')
#                             update_cm_yes = get_data('ctl00_ContentPlaceHolder1_RbVC2_0')
#                             dme_check_yes = get_data('ctl00_ContentPlaceHolder1_RbVC4_0')
#                             check_list = [update_family_yes,update_cm_yes,dme_check_yes]


#                             for word in teaching_list:
#                                 if word != 'checked':
#                                     visit_note_list.append('Care Check List not done '+ word)
#                                     break
#                             else:
#                                 pass

#                             back(2)
#                             print("Back 2 sleep B")
#                             sleep(20)
#                             print("did it go back to right page?")
#                             sound()
#                         visit_note_list.append(visit_date)
#                         print("VISIT NOTE LIST:",visit_note_list)
#                         print("FINDINGS",findings)


#                     collecting()
#                     append_to_excel(visit_note_list)
#                     print("loop done on one_page_visit")

#         one_page_visit_check()

# gather_visit_note_details()

# def check_meds():
#     click('//*[@id="ctl00_TreeView1t11"]')
#     random_xp = "//*[@id='ctl00_ContentPlaceHolder1_Panel1']"
#     find(random_xp)
#     medlist = '//*[@id="ctl00_ContentPlaceHolder1_GridView1"]'
#     allmedinfo = find(medlist)
#     allmed = allmedinfo.text
#     allmed = allmed.lower()

#     if 'frequency' in allmed:
#         pass
#     else:
#         findings.append("NO frequency in meds?")
#         print("NO frequency FOUND in meds")
#         print("ALLMED:",allmed)

#     if 'diet' in allmed:
#         pass
#     else:
#         print("NO DIET FOUND IN ORDERS")
#         findings.append("No diet found in meds.")
#         print("ALLMED:",allmed)


# for find in findings:
#     print(find)


# print('click_list = ',click_list)
# ############# now add it on excel###############

# #put this here so it would give me findings even if program crashed

# except Exception as e:
#     alert_sound()
#     print("Outer exception triggered")
#     print("EXCEPTION: ", e)
#     exc_type, exc_obj, exc_tb = sys.exc_info()
#     fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
#     print(exc_type, fname, exc_tb.tb_lineno)

#     for find in findings:
#         if find not in final_findings:
#             final_findings.append(find)

#     print("")
#     print("------------FINDINGS BELOW----------")
#     print("")
#     for final in final_findings:
#         print(final)


# ##### COLLECT MEDSS#############

# ##########REPEAT PROCESS FOR ALL SN VISIT NOTES###########


#             # add function to return finding if vital signs is not in normal range.


#          #for dementia
#          # make sure fast score is at least 7
#          #unable to ambulate w assistance
#          #unable to dress and bath w/o assistance
#          #urinary and fecal incontinence
#          #speak 6 or less words w no consistently meaningful communication

#             # Should have had one of these within past 12 months
#             #aspiration pneumonia,pyelonephritis,septicemia,decub ulcer multiple, stage 3-4,fever,recurrent after antibiotics,inability to maintain sufficient fluid, and calorie intake w 10% weight loss during previous six months or serum albumin 2.5gm

#             ## add frequency to findings.
#             #  IF narrative missing add to findings.

#             #make sure idg time is before idg time

#              # for date in date_object_list:
#               #      print(date)

#                 # if aohcs in narrative change to art od hospice


# # DID THEY HAVE ENOUGH VISITS BETWEEN IDG PER FREQUENCY?


# # make sure i have function that checks to make sure frequency is followed per orders

# #check make sure its long form

# # GATHER MISSING VISIT NOTES FROM REPORTS MISSING NOTES AND MESSAGE NURSE IF ITS NOT COMPLETED BY HOW LONG? 1 day?

# #make function if any vitals are out of range to check notes for it..

# # ana is responsible for making sure comprehensives done.


# print("FINDINGS: ",findings)


print('All Data')
for data in all_data:
    print(data)
