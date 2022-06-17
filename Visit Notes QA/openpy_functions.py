import openpyxl
from openpyxl import workbook, load_workbook
from openpyxl.styles import Alignment  # aligning
from openpyxl.styles import Font  # changing style
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
from openpyxl.styles.colors import Color
from openpyxl.styles import Font
import datetime

excel_location = 'C:/Users/johnc/Desktop/Visit Notes QA/QA VISIT NOTES.xlsx'
wb = load_workbook(excel_location)
ws = wb.active
max_row = ws.max_row


############################### ACCESS ####################################
def print_cell(letter_row):  # 'A1'
    print(ws[letter_row].value)

def get_row():
    for row in ws.iter_rows():
        row_num = str(row[0].row)
        # print("ROW NUM:",row_num)

def print_sheetnames():
    print(wb.sheetnames)

def read_column(letter):
    col_list = []
    for cell in ws[letter]:
        print(cell.value)
        col_list.append(cell.value)
    return col_list

def print_rows():
    for row in ws.iter_rows():
        print([cell.value for cell in row])

def get_last_row_num():
    last_row = ws.iter_rows()
    return last_row

def find_location_of_string(str_name):
    for row in ws.iter_rows(1):
        for cell in row:
            if cell.value == str_name:
                # print(ws.cell(row=cell.row, column=cell.column).value) #change column number for any cell value you want
                add = cell
                print(add)
                return add

def save_row_list():
    rows_list = []
    for row in ws.iter_rows():
        row_num = str(row[0].row)
        row = [cell.value for cell in row]
        row.append(row_num)
        rows_list.append(row)

def insert_new_column(letter):
    # inserts it before this one.
    ws.insert_cols(idx=letter)

def input_new_row():
    ws.insert_rows(7)

def delete_col():
    ws.delete_cols(6, 3)

def move_cells():
    ws.move_range("D4:F10", rows=-1, cols=2)

def convertTuple(tup):
    str = ''.join(tup)
    return str

def shorten_str_newlines(sentence):  #45 is good for excel
    sentence2 = sentence.split()
    words_length = 0
    word_index = 0
    for word in sentence2:
        length = len(word)
        words_length += length
        word_index += 1
        # print("WORDS LENGTH:",words_length)
        if words_length > 45:
            sentence2.insert(word_index, '\n')  
            words_length = 0 
        else:
            continue
    
    finished = ' '.join(sentence2) 
    # print(finished)
    return finished

####################### Create ADD ###############################

def create_worksheet(sheetname):
    names = wb.get_sheet_names()
    if sheetname not in names:
        wb.create_sheet(sheetname)
        wb.save(excel_name)
    else:
        print("THERE IS ALREADY A WORKSHEET WITH NAME:", sheetname)

    ws = wb[sheetname]

def add_list_to_row(list):
    ws.append(list)
    wb.save(excel_name)

def append_to_excel(list_to_add_in_new_row):
    ws.append(list_to_add_in_new_row)
    wb.save(excel_location)
# REQUIRES YOU TO HAVE CURRENT ROW BEFORE

def input_list_to_newline_excel(column, row, list_name):
    # counts the row to get the current row to put findings in right spot
    for row in ws.iter_rows(min_row=1, max_row=max_row):
        iterated_row = str(row[0].row)
        current_row = iterated_row
# for idg in idg_notes_list:

# NEED TO ITERATE EACH DIC TO IT
def apppend_list_of_dictionary_row(name_list, sheetname):
    key_list = []
    value_list = []
    names = wb.get_sheet_names()  # this ones deprecated
    # names = wb.sheetnames()

    if sheetname not in names:
        wb.create_sheet(sheetname)
        wb.save(excel_location)

    ws = wb[sheetname]
    for dictionary in name_list:
        try:
            for key, value in dictionary.items():
                if value is not None:
                    if value != '':
                        key_list.append(key)
                        value_list.append(value)
        except:
            pass

    # I CAN ADD NOT TO ADD IT IF THE DATE OF QA IS DONE ALREADY OR SOMETHING LIKE THAT>

    ws.append(key_list)
    ws.append(value_list)
    wb.save(excel_location)
    # print("Added Assessment to Excel")



def apppend_list_to_sheetname(list_name, sheetname):
    names = wb.get_sheet_names()  # this ones deprecated
    # names = wb.sheetnames()

    if sheetname not in names:
        wb.create_sheet(sheetname)
        wb.save(excel_location)

    ws = wb[sheetname]
    # I CAN ADD NOT TO ADD IT IF THE DATE OF QA IS DONE ALREADY OR SOMETHING LIKE THAT>

    ws.append(list_name)
    wb.save(excel_location)
    print("Added visit findings to Excel")

################# CHANGE #####################

def change_cell_value(letter_row, new_value):  # this didnt change btw.. only created new ones when empty?
    ws[letter_row] = new_value








                        ###     SPECIFIC FOR THIS PROGRAM     ###
##############################################################################################
parsed_findings = []  # ONLY USED IN 1 function dont delete for now

parsed_master_findings = []

## NOT SURE HOW IT DECIDES WHICH SHEET TO CHECK>> I"M GUESSING ITS THE LAST ACTIVE ONE

def append_list_of_dic_findings_excel(list_of_dic_name):  #Apends dic with key of 'row' referencing row and 'key sentence" to add to excel findings
    print("adding dic to row")
    for finding_dic in list_of_dic_name:
        for key, value in finding_dic.items():
            if key == 'row':
                row_location = finding_dic[key]
                # print("ROW LOC",row_location)
                location = 'A'+row_location 
                 #shorten sentence    
            sentence = finding_dic['sentence']
            shorter_sentence = shorten_str_newlines(sentence)

            change_cell_value(location, shorter_sentence)
            print("LOCATION:",location,shorter_sentence)

def save_rows_to_list():
    rows_list = []
    for row in ws.iter_rows():  # READS ALL ROWS PUTS ALL ROWS IN A LIST
        row_num = str(row[0].row)
        row = [cell.value for cell in row]
        row.append(row_num)
        rows_list.append(row)
        


    return rows_list

### ADDS ALL THESE TO FINDINGS..  Should  wait to add all of them.  These are under new_list.  its a list of dics with keys of row so i can append them to appropriate row in excel.
def find_rising_mac_scores():
    print("Starting find mac scores")

    def convertTuple(tup):  #PUT IT HERE TOO BECAUSE UnboundLocalError: local variable 'convertTuple' referenced before assignment
        str = ''.join(tup)
        return str

    ## NOTE I add a column at the bottom so this code might not work correctly again unless you account for it.
      # list of list


      #   check_mac_dates
    date_mac_name = []
    for row in rows_list:
        l = []
        visit_date = row[2]
        name = row[3]
        mac = row[10]
        num_row = row[-1]
        # print(visit_date,name,mac)
        l.append(visit_date)
        l.append(mac)
        l.append(name)
        l.append(num_row)
        date_mac_name.append(l)

    datetime_list = []
    # gets rid of data without data or mac
    for l in date_mac_name:
        # print("L",l)
        try:
            l2 = []
            date = l[0].split('/')
            month = int(date[0])
            day = int(date[1])
            remove_non_int = float((l[1]))
            year = int(date[2])
            l2.append(l[0])
            l2.append(l[1])
            l2.append(l[2])
            l2.append(l[-1])  # this is row num
            datetime_list.append(l2)
        except:
            pass

    newest_date = datetime.date.today()
    date_list = []
    for info in datetime_list:
        dic_info = {}
        dt = info[0]
        mac = info[1]
        name = info[2]
        row = info[3]
        dic_info['name'] = name
        dic_info['date'] = dt
        dic_info['mac'] = mac
        dic_info['row'] = row
        date_list.append(dic_info)
    # NOW PUT IT IN A DICTIONARY

    # checks date by order
    date_list.sort(key=lambda item: item['date'], reverse=False)

    print("SORTED DATE LIST")
    for date in date_list:
        print(date)


    previous_mac = 0
    mac_max_list = []
    for l in date_list:
        bad_dic = {}
        mac = l['mac']
        mac = float(mac)
        # print("MAC SCORE:",mac)

        mac_max_list.append(mac)
        max_mac_score = max(mac_max_list)
        min_mac_score = min(mac_max_list)
        
        add = False
        min_mac = False
        max_mac = False
        # print("MAC: ",mac, 'previous mac:',previous_mac, "MAX MAC:",max_mac_score)
        if previous_mac == 0:  #TO SEE IF IT WENT UP AFTER THE PREVIOUS ONE
            previous_mac = mac
        
        if mac > max_mac_score:         # checking if mac score went back up after it was charter lower  # STARTS WITH NEWEST ONE
            bad_dic['row'] = l['row']
            bad_dic['name'] = l['name']
            bad_dic['mac'] = l['mac']
            sentence = l['date'] + ' ' + l['name']+" charted higher Mac " + l['mac']+" from last visit." + ' was ' + str(previous_mac) # CHARTS WITH HIGHER MAC SCORES
            print("SENTENCE1:",sentence)
            bad_dic['sentence'] = sentence
            parsed_findings.append(bad_dic)
            print("Add to findings")

        if  mac > previous_mac:
            bad_dic['row'] = l['row']
            bad_dic['name'] = l['name']
            bad_dic['mac'] = l['mac']
            sentence = l['date'] + ' ' + l['name']+" charted higher Mac " + l['mac']+" from previous visit." + ' was ' + str(max_mac_score) # CHARTS WITH HIGHER MAC SCORES
            print("SENTENCE1:",sentence)
            bad_dic['sentence'] = sentence
            parsed_findings.append(bad_dic)

        if mac > min_mac_score:
            bad_dic['row'] = l['row']
            bad_dic['name'] = l['name']
            bad_dic['mac'] = l['mac']
            sentence = l['date'] + ' ' + l['name']+" charted higher Mac " + l['mac']+" from previous visit." + ' was ' + str(min_mac_score) # CHARTS WITH HIGHER MAC SCORES
            print("SENTENCE1:",sentence)
            bad_dic['sentence'] = sentence
            parsed_findings.append(bad_dic)

        print("MAC:",mac,"Previous mac:",previous_mac)
        print(previous_mac - mac)
        if  previous_mac - mac > 2:
            bad_dic['row'] = l['row']
            bad_dic['name'] = l['name']
            bad_dic['mac'] = l['mac']
            sentence = l['date'] + ' ' + l['name']+" mac dropped more than 3 points in a visit from previous visit?  They charted: " + l['mac']+ ' was ' + str(previous_mac) # CHARTS WITH HIGHER MAC SCORES
            print("SENTENCE1:",sentence)
            bad_dic['sentence'] = sentence
            parsed_findings.append(bad_dic)
   
  ######### WOR
                
        previous_mac = mac
    # NEED TO ADD SENTENCE TO THESE SENTENCES IN OLD ONES
    print("-----------Parsed FINDINGS------------")
    for finding in parsed_findings:
        print(finding)


    ## THIS IS GOOD NOW ADD IT ALL TO FINDINGS ##


    new_list = []
    for finding_dic in parsed_findings:  # ADDING SENTENCE TO THESE DICS
        new_dic = {}
        newest_sent = finding_dic['sentence'] 
        new_dic['row'] = finding_dic['row']
        new_dic['sentence'] = newest_sent
        new_list.append(new_dic)
        # print("NEW DIC",new_dic)
        # parsed_master_findings.append(new_dic)

    for item in new_list:
        # print("ITEMMMM",item)
        if item not in parsed_master_findings:
            parsed_master_findings.append(item)
        #APPEND EACH SENTENCE BY ROW TO column A
    return new_list


rows_list = save_rows_to_list()    
def check_decreasing_scores(check_row,row_name):  #WORKS FOR NUMBERS UNDER 6.. Can easily be modified
    cleaned_list = []
    ### Breaks apart the rows into list of lists # and filters them
    for row in rows_list:
        l = []
        name =  row[3]
        visit_date = row[2]
        notsure = row[13]  ## WHY?
        checkrow = row[check_row]   #THIS IS THE NUMBER
        # print("Checkrow:",checkrow)
        try:
            if int(checkrow) > 6:  # filters all non int
                # print("INT:",checkrow)
                continue
        except:
            continue
        # print("checkrow:",checkrow)
        num_row = row[-1]   #NUMBER IN THE ROW
        l.append(visit_date)
        l.append(checkrow)
        l.append(name)
        l.append(num_row)
        cleaned_list.append(l)
    # for l in cleaned_list:
    #     print("L",l)

    # for l in cleaned_list:                 ## #NEED TO SEE WHY I"M LOSING THE NUMS ABOVE
    #     print("L",l)
   
      ### PUTS THEM IN A list of dictionaries
    date_list = []
    for info in cleaned_list:
        dic_info = {}
        
        dt = info[0]
        number = info[1]
        name = info[2]
        
        row = info[3]
        dic_info['name'] = name
        dic_info['date'] = dt
        dic_info['number'] = number
        dic_info['row'] = row
        date_list.append(dic_info)
    # print("DATE LIST")
    # for date in date_list:
    #     print(date)
    # NOW PUT IT IN A DICTIONARY

    ######## SORTS THEM BY DATE
    date_list.sort(key=lambda item: item['date'], reverse=False)

    print("SORTED:")
    for date in date_list:
        print(date)
       ####
    initial_num = 0  # adds the numbers that went down according to date in a list
    lowest_dates = []
    previous_date = ''
    for l in date_list:
        bad_dic = {}
        number = l['number']
        name = l['name']
        date = l['date']
        row = l['row']
        if initial_num == 0:
            initial_num = number   # NEED TO SAVE ROW HERE SOMEWHERE TO DESIGNATE IT BACK

        

        if number == initial_num:
            continue

        if int(number) < int(initial_num):
            print("NUM: ",number,",initial_num: ",initial_num)
            print("Score lower from previous note ",l) # THESE ARE THE CHARTS WITH Previous LOWER MAC SCORES
            add_finding = l['name'] + ' '+ row_name + ' score was lower then previous visits:' + ' ' + str(number) + " /charted on: " + l['date']+ " was "+ initial_num + " previously."
            # print("ADD FINDING",add_finding)
            # Addd this lower charted mac before others  #NOW WHY ISNT IT BEING ADDED IN FINAL PARSED MASTER FINDINGS
            find_dic = {}
            find_dic['row'] = l['row']
            # print('find_dic',find_dic)
            find_dic['sentence'] = add_finding
            parsed_master_findings.append(find_dic)

    else:
        print("NO FINDINGS TO ADD")

## Get data from S and format them and append them to parsed_master_findings with row,date,
def add_findings_col_s():  ## MUST HAVE HEADER
    print("ADD FINDINGS COL S")

    for row in rows_list[2:]:
        col_s_findings = {}

        visit_date = row[2]
        name =  row[3]
        num_row = row[-1]
        visit_date = row[2]
        col_s = row[18]
        if col_s != None:
            if len(col_s) != 2:  # this got rid of empty ones
                if col_s != 'FINDINGS':
                    
                    col_s = col_s.replace('_',' ')
                    col_s = col_s.splitlines()
                    col_s = ','.join(col_s)
                    print("---------------------------------------------------------------------")   
                    print(len(col_s))
                    print("COL S:",col_s)
                    col_s_findings['row'] = num_row
                    col_s_findings['sentence'] = visit_date +  ' ' + name + " " + col_s
                    parsed_master_findings.append(col_s_findings)

 









### RUNNING CODE HERE

rows_list = save_rows_to_list()  #saves rows of excel to list of lists
find_rising_mac_scores()
check_decreasing_scores(11,'ambulation')  ## NOT 100% sure it adds dic to it
check_decreasing_scores(12,'toileting')
check_decreasing_scores(13,'transfer')
check_decreasing_scores(14,'dressing')
check_decreasing_scores(15,'feeding')
add_findings_col_s()


print("++++++++++++   parsed_master_findings   +++++++++++++++++++++")
for finding in parsed_master_findings:
    print(finding)

 
 
 #### ONCE I"VE COLLECTED ALL FINDINGS ####
 

#############  PARSE THROUGH IDG INFO< SC< SW INFO





# Then sort them by date to notify them of things needed changed. 


# THEN I NEED TO GATHER MED INFO
 
 






# ADD HEADING 
# insert_new_column(1)    ## ADDED TO COLUMN FOR FINDINGS IN COLUMN A
# append_list_of_dic_findings_excel(parsed_master_findings)



wb.save(excel_location)




